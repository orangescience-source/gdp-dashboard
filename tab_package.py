"""
Tab 6: 📦 업로드 패키지
확정된 모든 단계의 결과물을 하나의 패키지로 묶어 내보낸다.
"""
import io
from datetime import datetime

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from session_state_manager import (
    P1_CHANNEL, P1_TOPIC_TITLE, P1_CORE_MESSAGE, P1_EMOTION, P1_HOOK,
    P2_RESULT, P2_TITLE, P2_THUMBNAIL, P2_HOOK_30SEC, P2_IMAGE_PROMPT,
    P3_RESULT, P3_VIDEO_LENGTH, P3_STRUCTURE, P3_EMOTION_MAP, P3_MINI_HOOKS, P3_SCENE_META,
    P4_SCRIPT_FULL, P4_VIZ_MEMO, P4_CONFIRMED,
    render_pipeline_status,
    render_p1_confirmed_card,
    render_p2_confirmed_card,
    render_p3_confirmed_card,
    render_p4_confirmed_card,
)


# ──────────────────────────────────────────
# 완성도 체크리스트
# ──────────────────────────────────────────

def build_checklist() -> list:
    """각 단계별 완성 여부를 체크한다."""
    items = [
        ("채널 선택",           bool(st.session_state.get(P1_CHANNEL))),
        ("주제 확정 (P1)",      bool(st.session_state.get(P1_TOPIC_TITLE))),
        ("제목 확정 (P2)",      bool(st.session_state.get(P2_TITLE))),
        ("썸네일 문구 확정 (P2)", bool(st.session_state.get(P2_THUMBNAIL))),
        ("Hook 전략 확정 (P2)", bool(st.session_state.get(P2_HOOK_30SEC))),
        ("이미지 프롬프트 저장 (P2)", bool(st.session_state.get(P2_IMAGE_PROMPT))),
        ("대본 구조 설계 (P3)", bool(st.session_state.get(P3_STRUCTURE))),
        ("대본 작성 (P4)",      bool(st.session_state.get(P4_SCRIPT_FULL))),
        ("대본 최종 확정 (P4)", bool(st.session_state.get(P4_CONFIRMED))),
        ("시각화 메모 (P4→P5)",  bool(st.session_state.get(P4_VIZ_MEMO))),
    ]
    return items


# ──────────────────────────────────────────
# Excel 패키지 생성
# ──────────────────────────────────────────

def _hdr(cell, bg="4A90E2"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _dat(cell, bg="FFFFFF"):
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.fill = PatternFill("solid", fgColor=bg)


def export_package_excel() -> bytes:
    wb = Workbook()

    channel_name    = st.session_state.get(P1_CHANNEL, "")
    topic_title     = st.session_state.get(P1_TOPIC_TITLE, "")
    core_message    = st.session_state.get(P1_CORE_MESSAGE, "")
    target_emotion  = st.session_state.get(P1_EMOTION, "")
    hook_sentence   = st.session_state.get(P1_HOOK, "")
    title           = st.session_state.get(P2_TITLE, "")
    thumbnail       = st.session_state.get(P2_THUMBNAIL, "")
    hook_30sec      = st.session_state.get(P2_HOOK_30SEC, "")
    image_prompt    = st.session_state.get(P2_IMAGE_PROMPT, "")
    video_length    = st.session_state.get(P3_VIDEO_LENGTH, "")
    structure       = st.session_state.get(P3_STRUCTURE, {})
    emotion_map     = st.session_state.get(P3_EMOTION_MAP, [])
    mini_hooks      = st.session_state.get(P3_MINI_HOOKS, [])
    scene_meta      = st.session_state.get(P3_SCENE_META, {})
    script_full     = st.session_state.get(P4_SCRIPT_FULL, "")
    viz_memo        = st.session_state.get(P4_VIZ_MEMO, "")
    ts              = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── 시트 1: 기획 개요 ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📋 기획 개요"

    overview = [
        ("채널명",          channel_name),
        ("확정 주제",        topic_title),
        ("핵심 메시지",      core_message),
        ("타겟 감정",        target_emotion),
        ("Hook 문장",       hook_sentence),
        ("확정 제목",        title),
        ("썸네일 문구",      thumbnail),
        ("초반 30초 Hook",   hook_30sec),
        ("영상 길이",        video_length),
        ("패키지 생성일시",   ts),
    ]

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 70

    for i, (k, v) in enumerate(overview, 1):
        cell_k = ws1.cell(i, 1, k)
        cell_k.font = Font(bold=True)
        cell_k.fill = PatternFill("solid", fgColor="E8F0FE")
        cell_v = ws1.cell(i, 2, v)
        cell_v.alignment = Alignment(wrap_text=True, vertical="top")
        ws1.row_dimensions[i].height = 30 if "\n" not in str(v) else 60

    # ── 시트 2: 대본 구조 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("📐 대본 구조")
    section_order = ["hook", "teaser", "big_idea", "intro",
                     "body1", "body2", "body3", "body4",
                     "reveal", "impact", "end"]
    section_labels = {
        "hook": "HOOK", "teaser": "TEASER", "big_idea": "BIG IDEA", "intro": "INTRO",
        "body1": "BODY 1", "body2": "BODY 2", "body3": "BODY 3", "body4": "BODY 4",
        "reveal": "REVEAL", "impact": "IMPACT", "end": "END",
    }
    timecodes = {
        "hook": "00:00", "teaser": "01:00", "big_idea": "02:00", "intro": "03:00",
        "body1": "04:00", "body2": "07:00", "body3": "10:15", "body4": "13:30",
        "reveal": "17:00", "impact": "18:30", "end": "19:00",
    }

    ws2_headers = ["타임코드", "섹션", "정보 목적", "감정 목표", "장면 유형",
                   "주인공 역할", "보조 인물", "핵심 오브젝트", "리텐션 장치"]
    ws2.append(ws2_headers)
    for col in range(1, len(ws2_headers) + 1):
        _hdr(ws2.cell(1, col))
    ws2.row_dimensions[1].height = 28

    for idx, key in enumerate(section_order):
        sec = structure.get(key, {})
        bg = "F8F9FA" if idx % 2 == 0 else "FFFFFF"
        row = [
            timecodes.get(key, ""),
            section_labels.get(key, key.upper()),
            sec.get("info_purpose", ""),
            sec.get("emotion_goal", ""),
            sec.get("scene_type", ""),
            sec.get("protagonist_role", ""),
            sec.get("supporting_characters", ""),
            sec.get("key_objects", ""),
            sec.get("retention_device", ""),
        ]
        ws2.append(row)
        for col in range(1, len(row) + 1):
            _dat(ws2.cell(ws2.max_row, col), bg)
        ws2.row_dimensions[ws2.max_row].height = 45

    for i, w in enumerate([10, 12, 28, 20, 22, 22, 22, 22, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 3: 감정 지도 ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("🗺️ 감정 지도")
    em_headers = ["타임코드", "감정", "이모지", "설명"]
    ws3.append(em_headers)
    for col in range(1, len(em_headers) + 1):
        _hdr(ws3.cell(1, col), bg="7C4DFF")
    for idx, em in enumerate(emotion_map):
        bg = "F3E5F5" if idx % 2 == 0 else "FFFFFF"
        row = [em.get("timecode", ""), em.get("emotion", ""), em.get("emoji", ""), em.get("description", "")]
        ws3.append(row)
        for col in range(1, len(row) + 1):
            _dat(ws3.cell(ws3.max_row, col), bg)
    for i, w in enumerate([12, 18, 8, 50], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 4: 미니훅 ─────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("🪝 미니훅")
    mh_headers = ["타임코드", "미니훅 문장", "장면 전환 기능", "다음 장면 유형"]
    ws4.append(mh_headers)
    for col in range(1, len(mh_headers) + 1):
        _hdr(ws4.cell(1, col), bg="E53935")
    for idx, mh in enumerate(mini_hooks):
        bg = "FFF3E0" if idx % 2 == 0 else "FFFFFF"
        row = [mh.get("timecode", ""), mh.get("sentence", ""), mh.get("scene_transition", ""), mh.get("next_scene_type", "")]
        ws4.append(row)
        for col in range(1, len(row) + 1):
            _dat(ws4.cell(ws4.max_row, col), bg)
        ws4.row_dimensions[ws4.max_row].height = 40
    for i, w in enumerate([12, 40, 30, 25], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 5: 완성 대본 ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("📝 완성 대본")
    ws5.column_dimensions["A"].width = 100
    ws5["A1"] = f"채널: {channel_name}  |  제목: {title}  |  생성: {ts}"
    ws5["A1"].font = Font(bold=True)
    ws5["A2"] = ""

    if script_full:
        for line in script_full.split("\n"):
            ws5.append([line])
        ws5.column_dimensions["A"].width = 100
        for row in ws5.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="top")

    # ── 시트 6: 시각화 메모 ───────────────────────────────────────────────────
    if viz_memo:
        ws6 = wb.create_sheet("🎨 시각화 메모")
        ws6.column_dimensions["A"].width = 100
        ws6["A1"] = "시각화 연동 메모 (프롬프트 5용)"
        ws6["A1"].font = Font(bold=True, size=13)
        ws6["A2"] = ""
        for line in viz_memo.split("\n"):
            ws6.append([line])

    # ── 시트 7: 이미지 프롬프트 ─────────────────────────────────────────────
    if image_prompt:
        ws7 = wb.create_sheet("🖼️ 이미지 프롬프트")
        ws7.column_dimensions["A"].width = 120
        ws7["A1"] = "나노바나나 PRO 이미지 프롬프트"
        ws7["A1"].font = Font(bold=True, size=13)
        ws7["A2"] = ""
        ws7["A3"] = image_prompt
        ws7["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws7.row_dimensions[3].height = 200

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────
# 전체 TXT 패키지 생성
# ──────────────────────────────────────────

def export_package_txt() -> bytes:
    channel_name    = st.session_state.get(P1_CHANNEL, "")
    topic_title     = st.session_state.get(P1_TOPIC_TITLE, "")
    core_message    = st.session_state.get(P1_CORE_MESSAGE, "")
    target_emotion  = st.session_state.get(P1_EMOTION, "")
    title           = st.session_state.get(P2_TITLE, "")
    thumbnail       = st.session_state.get(P2_THUMBNAIL, "")
    hook_30sec      = st.session_state.get(P2_HOOK_30SEC, "")
    image_prompt    = st.session_state.get(P2_IMAGE_PROMPT, "")
    video_length    = st.session_state.get(P3_VIDEO_LENGTH, "")
    script_full     = st.session_state.get(P4_SCRIPT_FULL, "")
    viz_memo        = st.session_state.get(P4_VIZ_MEMO, "")
    ts              = datetime.now().strftime("%Y-%m-%d %H:%M")
    sep             = "━" * 50

    lines = [
        f"📦 유튜브 콘텐츠 업로드 패키지",
        f"채널: {channel_name}",
        f"생성일시: {ts}",
        sep,
        "",
        "【 1단계: 주제 기획 】",
        f"확정 주제: {topic_title}",
        f"핵심 메시지: {core_message}",
        f"타겟 감정: {target_emotion}",
        "",
        "【 2단계: 썸네일·제목 전략 】",
        f"확정 제목: {title}",
        f"썸네일 문구:\n{thumbnail}",
        f"초반 30초 Hook:\n{hook_30sec}",
        "",
    ]

    if image_prompt:
        lines += [
            "【 이미지 프롬프트 (나노바나나 PRO) 】",
            image_prompt,
            "",
        ]

    lines += [
        "【 3단계: 대본 구조 】",
        f"영상 길이: {video_length}",
        "(섹션별 구조 정보는 Excel 패키지 참고)",
        "",
        sep,
        "【 4단계: 완성 대본 】",
        "",
        script_full or "(대본 미작성)",
        "",
    ]

    if viz_memo:
        lines += [
            sep,
            "【 시각화 연동 메모 (프롬프트 5용) 】",
            "",
            viz_memo,
        ]

    return "\n".join(lines).encode("utf-8")


# ──────────────────────────────────────────
# 메인 탭 렌더링
# ──────────────────────────────────────────

def render_package_tab():
    render_pipeline_status()

    st.header("📦 업로드 패키지", divider="gray")
    st.caption(
        "확정된 모든 단계의 기획 결과물을 하나의 패키지로 묶어 내보냅니다.  \n"
        "유튜브 업로드 전 최종 체크리스트도 확인하세요."
    )

    # ── 완성도 체크리스트 ─────────────────────────────────────────────────────
    st.subheader("✅ 업로드 준비 체크리스트")
    checklist = build_checklist()
    done_count = sum(1 for _, v in checklist if v)
    total_count = len(checklist)

    progress_pct = done_count / total_count
    st.progress(progress_pct, text=f"완성도 {done_count}/{total_count}")

    col_check1, col_check2 = st.columns(2)
    for i, (label, done) in enumerate(checklist):
        icon = "✅" if done else "⬜"
        color = "#155724" if done else "#856404"
        bg = "#d4edda" if done else "#fff3cd"
        col = col_check1 if i % 2 == 0 else col_check2
        with col:
            st.markdown(
                f'<div style="background:{bg}; color:{color}; padding:6px 12px; '
                f'border-radius:8px; margin:3px 0; font-size:13px;">'
                f'{icon} {label}</div>',
                unsafe_allow_html=True,
            )

    all_done = done_count == total_count
    if all_done:
        st.success("🎉 모든 단계 완료! 패키지를 내보낼 준비가 되었습니다.")
    else:
        missing = [label for label, done in checklist if not done]
        st.warning(f"미완성 항목: {', '.join(missing)}")

    st.divider()

    # ── 확정 내용 카드 (각 단계) ─────────────────────────────────────────────
    st.subheader("📌 확정 내용 전체 확인")

    render_p1_confirmed_card(editable=False)
    render_p2_confirmed_card(editable=False)
    render_p3_confirmed_card(editable=False)
    render_p4_confirmed_card(editable=True)

    st.divider()

    # ── 빠른 요약 뷰 ─────────────────────────────────────────────────────────
    st.subheader("📋 빠른 요약")

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**채널 · 기획 정보**")
        ch = st.session_state.get(P1_CHANNEL, "—")
        tp = st.session_state.get(P1_TOPIC_TITLE, "—")
        ti = st.session_state.get(P2_TITLE, "—")
        em = st.session_state.get(P1_EMOTION, "—")
        vl = st.session_state.get(P3_VIDEO_LENGTH, "—")
        st.markdown(
            f"""
            | 항목 | 내용 |
            |------|------|
            | 채널 | {ch} |
            | 주제 | {tp} |
            | 제목 | {ti} |
            | 감정 | {em} |
            | 영상 길이 | {vl} |
            """
        )
    with col_b:
        script = st.session_state.get(P4_SCRIPT_FULL, "")
        structure = st.session_state.get(P3_STRUCTURE, {})
        mini_hooks = st.session_state.get(P3_MINI_HOOKS, [])
        viz_memo = st.session_state.get(P4_VIZ_MEMO, "")

        st.markdown("**대본 통계**")
        char_count = len(script)
        est_min = round(char_count / 500) if char_count else 0
        section_count = sum(1 for k in ["hook","teaser","big_idea","intro","body1","body2","body3","body4","reveal","impact","end"] if structure.get(k))
        memo_lines = len([l for l in viz_memo.split("\n") if l.strip()]) if viz_memo else 0
        st.markdown(
            f"""
            | 항목 | 값 |
            |------|------|
            | 총 글자 수 | {char_count:,}자 |
            | 예상 러닝타임 | 약 {est_min}분 |
            | 완성 섹션 | {section_count}/11 |
            | 미니훅 | {len(mini_hooks)}개 |
            | 시각화 메모 | {memo_lines}줄 |
            """
        )

    st.divider()

    # ── 업로드 준비 가이드 ────────────────────────────────────────────────────
    with st.expander("📋 유튜브 업로드 준비 가이드", expanded=False):
        st.markdown(
            """
            #### 업로드 전 최종 체크리스트

            **① 썸네일 제작**
            - 이미지 프롬프트를 나노바나나 PRO에 입력하여 이미지 생성
            - 생성된 이미지에 확정된 썸네일 문구 텍스트 오버레이
            - 최종 해상도: **1280×720px** (16:9), JPG/PNG
            - 유튜브 가이드라인 최종 검토 (선정성·폭력성 없는지 확인)

            **② 제목·설명 작성**
            - 확정 제목 그대로 사용 (SEO 키워드 포함 확인)
            - 설명란: 핵심 메시지 + 타임스탬프 + 링크 삽입

            **③ 대본 최종 검토**
            - 섹션별 흐름 재확인 (HOOK→TEASER→BIG IDEA→INTRO→BODY1-4→REVEAL→IMPACT→END)
            - [DATA_SKETCH_SCENE 후보] 태그 구간 시각화 자료 준비
            - 미니훅 위치(07:00/10:15/13:30/16:45) 편집 점 지정

            **④ 촬영·편집 준비**
            - 시각화 연동 메모를 기반으로 컷 리스트 작성
            - 와이드/미디엄/투샷/군중/사물 컷 비율 확인

            **⑤ 업로드 설정**
            - 카테고리: 주제에 맞게 선택
            - 태그: 제목 내 핵심 키워드 삽입
            - 공개 예약: 채널 시청자 활성 시간대 선택
            """
        )

    st.divider()

    # ── 내보내기 ──────────────────────────────────────────────────────────────
    st.subheader("📥 패키지 내보내기")

    channel_name = st.session_state.get(P1_CHANNEL, "채널")
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    ex_col1, ex_col2 = st.columns(2)

    with ex_col1:
        st.markdown("**📊 Excel 패키지 (전체)**")
        st.caption("기획 개요 · 대본 구조 · 감정 지도 · 미니훅 · 완성 대본 · 시각화 메모 · 이미지 프롬프트 (7시트)")
        try:
            excel_bytes = export_package_excel()
            st.download_button(
                "📥 Excel 패키지 다운로드",
                data=excel_bytes,
                file_name=f"업로드패키지_{channel_name}_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        except Exception as e:
            st.error(f"Excel 생성 오류: {e}")

    with ex_col2:
        st.markdown("**📄 TXT 패키지 (대본 전문)**")
        st.caption("전체 기획 정보 + 완성 대본 + 시각화 메모를 하나의 텍스트 파일로")
        try:
            txt_bytes = export_package_txt()
            st.download_button(
                "📥 TXT 패키지 다운로드",
                data=txt_bytes,
                file_name=f"업로드패키지_{channel_name}_{ts}.txt",
                mime="text/plain; charset=utf-8",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"TXT 생성 오류: {e}")

    # ── 이미지 프롬프트 빠른 복사 ────────────────────────────────────────────
    image_prompt = st.session_state.get(P2_IMAGE_PROMPT, "")
    if image_prompt:
        st.divider()
        st.subheader("🖼️ 이미지 프롬프트 (나노바나나 PRO)")
        st.caption("아래 프롬프트를 복사하여 나노바나나에 바로 붙여넣으세요.")
        st.text_area(
            "이미지 프롬프트",
            value=image_prompt,
            height=150,
            key="pkg_image_prompt",
        )

    # ── 최종 메시지 ───────────────────────────────────────────────────────────
    st.divider()
    if all_done:
        st.success(
            "🎬 **업로드 패키지 완성!**\n\n"
            "모든 단계가 완료되었습니다.  \n"
            "Excel/TXT 파일을 다운로드하고 썸네일 제작을 시작하세요! 🚀"
        )
    else:
        st.info(
            "💡 미완성 단계를 완료한 후 패키지를 다시 생성하세요.  \n"
            "미완성 항목이 있어도 현재까지 완성된 내용으로 내보내기가 가능합니다."
        )
