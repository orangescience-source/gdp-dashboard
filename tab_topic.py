import os
import io
import json
from datetime import datetime

import streamlit as st
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

from channel_db import CHANNEL_DB
from prompts import PROMPT_1_SYSTEM
from session_state_manager import (
    P1_CHANNEL, P1_BENCHMARK, P1_TOPIC_TITLE, P1_CORE_MESSAGE, P1_EMOTION, P1_HOOK,
    render_pipeline_status,
)
from youtube_researcher import (
    search_trending_videos,
    format_view_count,
    SEARCH_CRITERIA,
)

CHANNEL_NAMES = list(CHANNEL_DB.keys())


def _get_client():
    """API 키 우선순위: st.secrets → 환경변수"""
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        key = ""
    key = key or os.environ.get("ANTHROPIC_API_KEY", "")
    return anthropic.Anthropic(api_key=key)


def build_persona_block(channel_name: str) -> str:
    if channel_name not in CHANNEL_DB:
        return "채널 정보 없음 — 중립적 분석가 페르소나로 진행"
    info = CHANNEL_DB[channel_name]
    return (
        f"\n채널명: {channel_name}"
        f"\n주인공: {info['host']} ({info['host_desc']})"
        f"\n색상 정체성: {info['color_primary']} / {info['color_secondary']}"
        f"\n톤앤매너: {info['tone']}"
        f"\n타겟 오디언스: {info['target']}"
        f"\nSEO 핵심 키워드: {', '.join(info['seo_keywords'])}"
        f"\n썸네일 전략: {info['thumbnail_style']}"
        f"\n시각 무드: {info['visual_mood']}\n"
    )


def _extract_json(raw: str) -> dict:
    """응답 텍스트에서 JSON 객체를 추출하고 파싱한다."""
    text = raw.strip()

    # 마크다운 코드펜스 제거
    if text.startswith("```"):
        parts = text.split("```")
        for part in parts:
            part = part.strip()
            if part.startswith("json"):
                part = part[4:].strip()
            try:
                return json.loads(part)
            except json.JSONDecodeError:
                continue

    # 전체 텍스트 직접 파싱 시도
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 중괄호로 시작하는 첫 번째 JSON 객체 추출
    start = text.find("{")
    if start == -1:
        raise json.JSONDecodeError("No JSON object found", text, 0)

    # 중첩 중괄호 균형을 맞춰 끝 위치 탐색
    depth = 0
    end = -1
    in_string = False
    escape = False
    for i, ch in enumerate(text[start:], start):
        if escape:
            escape = False
            continue
        if ch == "\\" and in_string:
            escape = True
            continue
        if ch == '"':
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i
                break

    if end == -1:
        # 응답이 잘린 경우 — 끝까지 있는 텍스트로 파싱 시도
        raise json.JSONDecodeError(
            "JSON이 중간에 잘렸습니다 (응답 토큰 초과)", text, len(text)
        )

    json_str = text[start: end + 1]
    return json.loads(json_str)


def call_claude_prompt1(channel_name, benchmark_input, video_length, extra_req,
                        planning_context=""):
    persona_block = build_persona_block(channel_name)
    system_prompt = PROMPT_1_SYSTEM.format(persona_block=persona_block)

    planning_section = ""
    if planning_context.strip():
        planning_section = (
            f"\n[사전 기획 정보 - 반드시 반영]\n"
            f"{planning_context}\n"
            "위 기획 정보를 최대한 반영하여 주제를 발굴해주세요.\n"
        )

    user_message = (
        f"채널명: {channel_name}\n"
        f"벤치마킹 대상: {benchmark_input}\n"
        f"원하는 영상 길이: {video_length}\n"
        f"추가 요구사항: {extra_req if extra_req else '없음'}\n"
        f"{planning_section}\n"
        "위 정보를 바탕으로 프롬프트 1을 실행하여 JSON 형식으로 결과를 반환하라.\n"
        "중요: 각 필드는 간결하게 작성하라 (한 필드당 100자 이내). "
        "응답은 반드시 완전한 JSON이어야 한다."
    )
    client = _get_client()

    last_error = None
    for attempt in range(2):  # 실패 시 1회 재시도
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=16000,
            system=system_prompt,
            messages=[{"role": "user", "content": user_message}],
        )
        raw = response.content[0].text.strip()
        try:
            return _extract_json(raw)
        except json.JSONDecodeError as e:
            last_error = e
            if attempt == 0:
                continue  # 한 번 더 시도

    raise json.JSONDecodeError(
        f"2회 시도 후 파싱 실패: {last_error}",
        "", 0
    )


# ── UI 컴포넌트 ───────────────────────────────────────────────────────────────

def render_persona_card(channel_name: str):
    if channel_name not in CHANNEL_DB:
        st.warning("채널 정보를 찾을 수 없습니다. 직접 입력한 채널명으로 진행합니다.")
        return
    info = CHANNEL_DB[channel_name]
    color = info["color_primary"]
    st.markdown(
        f"""
        <div style="
            border-left: 4px solid {color};
            padding: 12px 16px;
            background: #f8f9fa;
            border-radius: 0 8px 8px 0;
            margin-bottom: 16px;
            color: #1a1a1a;
        ">
            <div style="font-size:12px; color:#666; margin-bottom:4px;">자동 매칭된 채널 페르소나</div>
            <div style="font-size:16px; font-weight:700; color:#111;">{channel_name}</div>
            <div style="font-size:13px; color:#333; margin-top:6px; line-height:1.6;">
                🎭 <b>{info['host']}</b> — {info['host_desc']}<br>
                🎯 {info['tone']}<br>
                👥 타겟: {info['target']}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_methodology_panel(methodology: dict):
    """분석 방법론 및 신뢰도 패널 렌더링"""
    if not methodology:
        return

    conf = methodology.get("confidence_level", "")
    conf_color = {"높음": "#2ecc71", "중간": "#f39c12", "낮음": "#e74c3c"}.get(conf, "#888")
    conf_icon = {"높음": "🟢", "중간": "🟡", "낮음": "🔴"}.get(conf, "⚪")

    patterns = methodology.get("key_patterns_found", [])
    patterns_html = "".join(
        f'<li style="margin-bottom:3px;">{p}</li>' for p in patterns
    )

    score_criteria = methodology.get("score_criteria", {})

    st.markdown(
        f"""
        <div style="
            border: 1px solid #dce3ed;
            border-radius: 12px;
            padding: 16px 20px;
            background: #f7faff;
            margin-bottom: 16px;
        ">
            <div style="font-size:14px; font-weight:700; color:#1a1a1a; margin-bottom:10px;">
                🔬 분석 로직 & 신뢰도
            </div>

            <div style="display:flex; gap:12px; flex-wrap:wrap; margin-bottom:12px;">
                <div style="background:#fff; border:1px solid #ddd; border-radius:8px;
                            padding:8px 14px; flex:1; min-width:140px;">
                    <div style="font-size:11px; color:#888;">신뢰도</div>
                    <div style="font-size:15px; font-weight:700; color:{conf_color};">
                        {conf_icon} {conf}
                    </div>
                    <div style="font-size:11px; color:#666; margin-top:3px;">
                        {methodology.get("confidence_reason", "")}
                    </div>
                </div>
                <div style="background:#fff; border:1px solid #ddd; border-radius:8px;
                            padding:8px 14px; flex:1; min-width:140px;">
                    <div style="font-size:11px; color:#888;">분석 유형</div>
                    <div style="font-size:13px; font-weight:600; color:#333; margin-top:2px;">
                        {methodology.get("benchmark_type", "")}
                    </div>
                </div>
            </div>

            <div style="font-size:12px; color:#555; background:#fff3cd; border-radius:6px;
                        padding:8px 12px; margin-bottom:10px; border-left:3px solid #f39c12;">
                ⚠️ <b>AI 분석 기준:</b> {methodology.get("analysis_basis", "")}
            </div>

            <div style="margin-bottom:10px;">
                <div style="font-size:12px; font-weight:600; color:#333; margin-bottom:4px;">
                    📌 발견된 핵심 패턴
                </div>
                <ul style="margin:0; padding-left:18px; font-size:12px; color:#555;">
                    {patterns_html}
                </ul>
            </div>

            <div style="font-size:12px; color:#555; background:#eef5ff; border-radius:6px;
                        padding:8px 12px; margin-bottom:10px;">
                📊 <b>시장 맥락:</b> {methodology.get("market_context", "")}
            </div>

            <details style="font-size:12px; color:#555;">
                <summary style="cursor:pointer; font-weight:600; color:#444; margin-bottom:6px;">
                    📐 점수 산정 기준 보기
                </summary>
                <div style="padding:8px 0 0 4px;">
                    <div style="margin-bottom:4px;">
                        <b>검색량:</b> {score_criteria.get("search_volume", "")}
                    </div>
                    <div style="margin-bottom:4px;">
                        <b>경쟁도:</b> {score_criteria.get("competition", "")}
                    </div>
                    <div style="margin-bottom:4px;">
                        <b>CTR:</b> {score_criteria.get("ctr", "")}
                    </div>
                    <div>
                        <b>페르소나 적합도:</b> {score_criteria.get("persona_fit", "")}
                    </div>
                </div>
            </details>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_topic_card(topic: dict, is_top: bool = False):
    rank = topic.get("rank", 0)
    fit = int(topic.get("persona_fit", 3))
    stars = "⭐" * fit + "☆" * (5 - fit)

    ctr = topic.get("expected_ctr", "")
    if "8%" in ctr:
        ctr_color = "#2ecc71"
    elif "5-8" in ctr or "5%" in ctr:
        ctr_color = "#f39c12"
    else:
        ctr_color = "#e74c3c"

    badge_map = {"높음": "🔴", "중간": "🟡", "낮음": "🟢"}
    diff_map = {"하": "🟢", "중": "🟡", "상": "🔴"}
    search_badge = badge_map.get(topic.get("search_volume", ""), "")
    comp_badge = badge_map.get(topic.get("competition", ""), "")
    diff_badge = diff_map.get(topic.get("difficulty", ""), "")

    border_style = "2px solid #f39c12" if is_top else "1px solid #e0e0e0"
    header_bg = "#fff9e6" if is_top else "#ffffff"
    top_badge = "🏆 최우선 추천  " if is_top else ""

    st.markdown(
        f"""
        <div style="
            border: {border_style};
            border-radius: 12px;
            padding: 16px;
            margin-bottom: 12px;
            background: {header_bg};
            color: #1a1a1a;
        ">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                <span style="font-size:13px; color:#555; font-weight:600;">{top_badge}주제 {rank}</span>
                <span style="font-size:13px;">{stars}</span>
            </div>
            <div style="font-size:17px; font-weight:700; color:#111; margin-bottom:8px;
                        line-height:1.4;">{topic.get("title","")}</div>
            <div style="font-size:14px; color:#333; margin-bottom:10px; line-height:1.5;">
                {topic.get("core_message","")}
            </div>
            <div style="display:flex; gap:8px; flex-wrap:wrap; margin-bottom:10px;">
                <span style="font-size:12px; color:#1a237e; background:#e8eaf6;
                             padding:4px 10px; border-radius:20px; font-weight:500;">
                    검색량 {search_badge} {topic.get("search_volume","")}
                </span>
                <span style="font-size:12px; color:#4a148c; background:#f3e5f5;
                             padding:4px 10px; border-radius:20px; font-weight:500;">
                    경쟁도 {comp_badge} {topic.get("competition","")}
                </span>
                <span style="font-size:12px; color:#1b5e20; background:#e8f5e9;
                             padding:4px 10px; border-radius:20px; font-weight:500;">
                    난이도 {diff_badge} {topic.get("difficulty","")}
                </span>
                <span style="font-size:12px; color:{ctr_color}; background:#f8fff8;
                             padding:4px 10px; border-radius:20px; border:1px solid {ctr_color};
                             font-weight:600;">
                    CTR {ctr}
                </span>
            </div>
            <div style="font-size:13px; color:#333; background:#f5f5f5;
                        padding:10px 14px; border-radius:8px; margin-bottom:8px;
                        border-left:3px solid #bbb;">
                💡 <b>차별화:</b> {topic.get("differentiation","")}
            </div>
            <div style="font-size:13px; color:#333; background:#e8f4fd;
                        padding:10px 14px; border-radius:8px;
                        border-left:3px solid #4A90E2;">
                🎭 <b>채널 앵글:</b> {topic.get("channel_angle","")}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.expander(f"주제 {rank} 상세 보기"):
        st.markdown(f"**타겟 감정:** {topic.get('target_emotion','')}")
        st.markdown(f"**페르소나 적합 이유:** {topic.get('persona_fit_reason','')}")
        st.markdown(f"**필요 리서치:** {topic.get('research_needed','')}")
        st.markdown(f"**예상 제작 시간:** {topic.get('production_time','')}")
        st.info(f"🎤 Hook 문장: \"{topic.get('hook_sentence','')}\"")

        # 분석 근거 (reasoning) 섹션
        reasoning = topic.get("reasoning", {})
        if reasoning:
            st.markdown("---")
            st.markdown("##### 🧠 이 주제 선정 근거 (AI 분석 로직)")
            r_col1, r_col2 = st.columns(2)
            with r_col1:
                st.markdown(
                    f"""
                    <div style="background:#f0f8ff; border-radius:8px; padding:10px 12px; margin-bottom:8px; font-size:12px;">
                        <div style="font-weight:700; color:#1a6bbf; margin-bottom:4px;">💡 선정 이유</div>
                        <div style="color:#333;">{reasoning.get("why_selected","")}</div>
                    </div>
                    <div style="background:#f5fff5; border-radius:8px; padding:10px 12px; margin-bottom:8px; font-size:12px;">
                        <div style="font-weight:700; color:#27ae60; margin-bottom:4px;">📈 검색량 판단 근거</div>
                        <div style="color:#333;">{reasoning.get("search_volume_basis","")}</div>
                    </div>
                    <div style="background:#fffaf0; border-radius:8px; padding:10px 12px; font-size:12px;">
                        <div style="font-weight:700; color:#e67e22; margin-bottom:4px;">⚔️ 경쟁도 판단 근거</div>
                        <div style="color:#333;">{reasoning.get("competition_basis","")}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            with r_col2:
                st.markdown(
                    f"""
                    <div style="background:#fdf5ff; border-radius:8px; padding:10px 12px; margin-bottom:8px; font-size:12px;">
                        <div style="font-weight:700; color:#8e44ad; margin-bottom:4px;">🖱️ CTR 예측 근거</div>
                        <div style="color:#333;">{reasoning.get("ctr_basis","")}</div>
                    </div>
                    <div style="background:#f0faff; border-radius:8px; padding:10px 12px; margin-bottom:8px; font-size:12px;">
                        <div style="font-weight:700; color:#2980b9; margin-bottom:4px;">🎭 페르소나 적합 이유</div>
                        <div style="color:#333;">{reasoning.get("persona_basis","")}</div>
                    </div>
                    <div style="background:#fff5f5; border-radius:8px; padding:10px 12px; font-size:12px; border-left:3px solid #e74c3c;">
                        <div style="font-weight:700; color:#e74c3c; margin-bottom:4px;">⚠️ 리스크 / 주의사항</div>
                        <div style="color:#333;">{reasoning.get("risk","")}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        if st.button(
            "이 주제로 썸네일·제목 전략 생성 →",
            key=f"to_prompt2_{rank}",
            type="primary",
        ):
            st.session_state[P1_CHANNEL]      = st.session_state.get("p1_channel", "")
            st.session_state[P1_TOPIC_TITLE]  = topic.get("title", "")
            st.session_state[P1_CORE_MESSAGE] = topic.get("core_message", "")
            st.session_state[P1_EMOTION]      = topic.get("target_emotion", "")
            st.session_state[P1_HOOK]         = topic.get("hook_sentence", "")
            st.success("✅ 주제가 저장되었습니다! 상단 '🎨 썸네일·제목' 탭으로 이동하세요.")


# ── Excel 내보내기 ────────────────────────────────────────────────────────────

def export_to_excel(result: dict, channel_name: str, benchmark_input: str) -> bytes:
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    top_fill = PatternFill("solid", fgColor="FFF9E6")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 시트 1: 분석 요약 ──
    ws1 = wb.active
    ws1.title = "분석 요약"
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 55

    summary_rows = [
        ("채널명", channel_name),
        ("분석 일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("벤치마킹 대상", benchmark_input),
        ("", ""),
        ("── 최우선 추천 주제 ──", ""),
        ("추천 순위", str(result.get("top_pick", {}).get("rank", ""))),
        ("선정 이유", result.get("top_pick", {}).get("reason", "")),
        ("첫 24시간 예상 조회수", result.get("top_pick", {}).get("first_24h_views", "")),
        ("7일 누적 예상", result.get("top_pick", {}).get("day7_views", "")),
        ("구독 전환율", result.get("top_pick", {}).get("subscribe_rate", "")),
        ("", ""),
        ("── SEO 키워드 ──", ""),
        ("메인 키워드", ", ".join(result.get("seo", {}).get("main_keywords", []))),
        ("롱테일 키워드", ", ".join(result.get("seo", {}).get("longtail_keywords", []))),
        ("해시태그", " ".join(result.get("seo", {}).get("hashtags", []))),
    ]
    for i, (k, v) in enumerate(summary_rows, 1):
        c_k = ws1.cell(i, 1, k)
        c_v = ws1.cell(i, 2, v)
        c_k.font = Font(bold=True)
        c_v.border = border

    ws1.cell(19, 1, "").value = None
    # 주제 점수 비교 표
    score_headers = ["순위", "제목", "검색량", "경쟁도", "CTR", "적합도", "난이도"]
    for col, h in enumerate(score_headers, 1):
        c = ws1.cell(20, col, h)
        c.fill = header_fill
        c.font = header_font
        c.border = border

    top_rank = result.get("top_pick", {}).get("rank", 0)
    for i, topic in enumerate(result.get("topics", []), 21):
        is_top = topic.get("rank") == top_rank
        vals = [
            topic.get("rank"), topic.get("title"),
            topic.get("search_volume"), topic.get("competition"),
            topic.get("expected_ctr"),
            "⭐" * int(topic.get("persona_fit", 0)),
            topic.get("difficulty"),
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(i, col, str(val) if val is not None else "")
            if is_top:
                c.fill = top_fill
            c.border = border

    # ── 시트 2: 상세 분석 ──
    ws2 = wb.create_sheet("상세 분석")
    headers2 = [
        "순위", "제목", "핵심 메시지", "타겟 감정",
        "차별화", "채널 앵글", "검색량", "경쟁도", "CTR",
        "적합도", "필요 리서치", "제작 시간", "난이도",
    ]
    col_widths2 = [6, 32, 40, 20, 35, 35, 10, 10, 10, 8, 35, 14, 8]
    for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        c = ws2.cell(1, col, h)
        c.fill = header_fill
        c.font = header_font
        c.border = border

    for i, topic in enumerate(result.get("topics", []), 2):
        row_data = [
            topic.get("rank"), topic.get("title"), topic.get("core_message"),
            topic.get("target_emotion"), topic.get("differentiation"),
            topic.get("channel_angle"), topic.get("search_volume"),
            topic.get("competition"), topic.get("expected_ctr"),
            topic.get("persona_fit"), topic.get("research_needed"),
            topic.get("production_time"), topic.get("difficulty"),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(i, col, str(val) if val is not None else "")
            c.border = border
        ws2.row_dimensions[i].height = 48

    # ── 시트 3: Hook 문장 모음 ──
    ws3 = wb.create_sheet("Hook 문장")
    for col, (h, w) in enumerate(zip(["순위", "제목", "Hook 문장"], [6, 32, 65]), 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        c = ws3.cell(1, col, h)
        c.fill = header_fill
        c.font = header_font
        c.border = border

    for i, topic in enumerate(result.get("topics", []), 2):
        ws3.cell(i, 1, topic.get("rank")).border = border
        ws3.cell(i, 2, topic.get("title", "")).border = border
        c = ws3.cell(i, 3, topic.get("hook_sentence", ""))
        c.border = border
        ws3.row_dimensions[i].height = 36

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 메인 탭 함수 ──────────────────────────────────────────────────────────────

def render_topic_tab():
    render_pipeline_status()

    # Stage 1 자동 입력 안내 배너
    if st.session_state.get("_autofill_stage1_applied"):
        _s1_src = st.session_state.get("_autofill_stage1_source", "")
        _s1_info_col, _s1_dismiss_col = st.columns([5, 1])
        with _s1_info_col:
            st.info(
                f"🔗 **이전 단계(니치 발굴)에서 자동 입력됨** — 니치: '{_s1_src}'\n\n"
                "벤치마킹 대상, 영상 주제, 핵심 메시지가 자동으로 채워져 있습니다. "
                "확인 후 필요시 수정하세요."
            )
        with _s1_dismiss_col:
            if st.button("✕ 닫기", key="_dismiss_stage1_banner"):
                st.session_state["_autofill_stage1_applied"] = False
                st.rerun()

    st.header("📊 트렌드 기반 주제 발굴 & 경쟁 분석기")
    st.caption("Claude AI가 채널 페르소나에 맞게 차별화 주제 5가지를 도출합니다.")

    # ── 채널 선택 (전체 너비) ─────────────────────────────────────────────────
    channel_options = ["직접 입력"] + CHANNEL_NAMES
    channel_select = st.selectbox(
        "채널 선택",
        channel_options,
        key="p1_channel_select",
        help="13개 채널 중 선택하거나 직접 입력하세요.",
    )
    if channel_select == "직접 입력":
        channel_name = st.text_input("채널명 직접 입력", key="p1_channel_custom")
    else:
        channel_name = channel_select
    # 항상 session_state에 저장 (직접 입력 포함)
    if channel_name:
        st.session_state[P1_CHANNEL] = channel_name

    if channel_name and channel_name in CHANNEL_DB:
        render_persona_card(channel_name)

    # ── 원터치 벤치마킹 주제 발굴 ────────────────────────────────────────────
    st.divider()
    st.subheader("🎯 원터치 벤치마킹 주제 발굴")
    st.caption(
        f"선택한 채널 페르소나에 맞는 "
        f"최근 {SEARCH_CRITERIA['period_days']}일 내 "
        f"인기 영상을 자동으로 검색합니다. "
        f"(최소 조회수 "
        f"{format_view_count(SEARCH_CRITERIA['min_views'])} 이상)"
    )

    col_btn1, col_btn2 = st.columns([2, 1])

    with col_btn2:
        max_topics = st.selectbox(
            "발굴 개수",
            options=[3, 5, 10],
            index=1,
            key="yt_max_topics"
        )

    with col_btn1:
        yt_search_btn = st.button(
            "🔍 YouTube 인기 영상 자동 발굴",
            type="primary",
            use_container_width=True,
            disabled=not st.session_state.get(
                "p1_channel", ""
            ),
            key="yt_search_btn"
        )

    # YouTube API 키 미설정 안내
    if "YOUTUBE_API_KEY" not in st.secrets:
        st.warning(
            "⚠️ YouTube API 키가 설정되지 않았습니다. "
            "Streamlit secrets에 YOUTUBE_API_KEY 를 추가해주세요."
        )

    # 발굴 버튼 클릭 처리
    if yt_search_btn:
        channel = st.session_state.get("p1_channel", "")
        with st.spinner(
            f"🔍 {channel} 채널에 맞는 인기 영상을 검색 중..."
        ):
            videos = search_trending_videos(
                channel_name=channel,
                max_topics=max_topics
            )

        if not videos:
            st.error(
                "검색 결과가 없습니다. "
                "잠시 후 다시 시도하거나 YouTube API 키를 확인해주세요."
            )
        else:
            st.session_state["yt_search_results"] = videos
            st.success(
                f"✅ {len(videos)}개 인기 영상 발굴 완료!"
            )

    # 검색 결과 표시
    videos = st.session_state.get("yt_search_results", [])
    if videos:
        st.markdown("#### 📊 발굴된 벤치마킹 후보")
        st.caption(
            "아래 영상을 클릭하면 벤치마킹 대상으로 자동 입력됩니다."
        )

        for i, video in enumerate(videos):
            with st.expander(
                f"{'🥇' if i==0 else '🥈' if i==1 else '🥉' if i==2 else '📌'} "
                f"{video['title'][:45]}... "
                f"| {format_view_count(video['view_count'])}회",
                expanded=(i == 0)
            ):
                col_a, col_b, col_c = st.columns(3)
                col_a.metric(
                    "총 조회수",
                    format_view_count(video["view_count"])
                )
                col_b.metric(
                    "일평균 조회수",
                    format_view_count(video["daily_views"])
                )
                col_c.metric(
                    "업로드 후",
                    f"{video['days_since']}일"
                )

                st.markdown(
                    f"**채널:** {video['channel']} | "
                    f"**업로드:** {video['published_at']} | "
                    f"**검색 키워드:** #{video['keyword']}"
                )
                st.markdown(
                    f"🔗 [영상 바로가기]({video['url']})"
                )

                if st.button(
                    "✅ 이 영상으로 벤치마킹 시작",
                    key=f"select_video_{i}",
                    use_container_width=True,
                    type="secondary"
                ):
                    # 벤치마킹 입력창에 URL 자동 입력
                    st.session_state["benchmark_input"]    = video["url"]
                    st.session_state["benchmark_title"]    = video["title"]
                    st.session_state["p1_benchmark_input"] = video["url"]
                    st.success(
                        f"✅ '{video['title'][:30]}...' 선택됨! "
                        "아래 주제 발굴 섹션에서 생성 버튼을 눌러주세요."
                    )
                    st.rerun()

        # Stage 2: 최상위 영상 자동 적용
        st.divider()
        _s2_col, _s2_btn_col = st.columns([3, 1])
        with _s2_col:
            _apply_stage2 = st.checkbox(
                "상위 영상을 벤치마킹 대상에 자동으로 적용",
                value=False,
                key="_stage2_checkbox",
            )
        with _s2_btn_col:
            _stage2_btn = st.button(
                "🔍 자동 적용 →",
                use_container_width=True,
                disabled=not _apply_stage2,
                key="_stage2_apply_btn",
            )
        if _stage2_btn and _apply_stage2 and videos:
            _top_video = videos[0]
            st.session_state["p1_benchmark_input"] = _top_video["url"]
            st.session_state["benchmark_input"] = _top_video["url"]
            st.session_state["benchmark_title"] = _top_video["title"]
            st.session_state["_autofill_stage2_applied"] = True
            st.session_state["_autofill_stage2_source"] = _top_video["title"]
            st.rerun()

    st.divider()

    # ── 기획 입력 폼 ────────────────────────────────────────────────────────────
    st.subheader("📋 기획 정보 입력")
    st.caption(
        "입력할수록 더 정확한 주제가 발굴됩니다. "
        "필수 항목만 입력해도 됩니다."
    )

    stage_hooks = [
        ("STAGE 1 HOOK",        "강렬한 첫 장면"),
        ("STAGE 2 PROBLEM",     "문제 제기"),
        ("STAGE 3 CONTEXT",     "배경 설명"),
        ("STAGE 4 TWIST",       "반전/핵심 폭로"),
        ("STAGE 5 DEEP DIVE",   "심층 분석"),
        ("STAGE 6 IMPLICATION", "영향/결과"),
        ("STAGE 7 ACTION",      "행동 제안"),
        ("STAGE 8 END",         "마무리 + 예고"),
    ]

    with st.expander("📋 기획 정보 입력 (클릭하여 펼치기)", expanded=True):
        # 자동 입력 안내
        _autofill_flags = [
            ("_autofill_stage1_applied", "_autofill_stage1_source", "니치 발굴"),
            ("_autofill_stage2_applied", "_autofill_stage2_source", "YouTube 영상"),
            ("_autofill_stage3_applied", "_autofill_stage3_source", "AI 분석 결과"),
        ]
        _active_sources = [
            f"{label} ({st.session_state.get(src_key, '')[:20]})"
            for flag_key, src_key, label in _autofill_flags
            if st.session_state.get(flag_key)
        ]
        if _active_sources:
            st.info(
                f"🔗 **이전 단계에서 자동 입력됨** — 출처: {', '.join(_active_sources)}\n\n"
                "값을 직접 수정할 수 있습니다."
            )

        col_a, col_b = st.columns(2)

        with col_a:
            st.text_input(
                "1️⃣ 영상 주제 (선택)",
                placeholder='예: "한국 부동산 폭락의 진실"',
                key="planning_topic",
                help="구체적인 주제가 있으면 입력. 없으면 비워두세요.",
            )
            st.text_input(
                "2️⃣ 핵심 메시지 (선택)",
                placeholder="시청자가 가져갈 1문장",
                key="planning_message",
                help="예: 지금 집을 사면 안 되는 이유 3가지",
            )
            st.text_input(
                "3️⃣ 타겟 감정 흐름 (선택)",
                placeholder="예: 공포 → 분노 → 희망",
                key="planning_emotion",
                help="시청자가 느껴야 할 감정 흐름",
            )
            st.text_area(
                "6️⃣ 주요 수치/데이터 (선택)",
                placeholder=(
                    "예: 서울 아파트 평균가 12억\n"
                    "금리 3.5% → 5.0% 상승\n"
                    "거래량 전년 대비 40% 감소"
                ),
                height=100,
                key="planning_data",
                help="대본에 반영할 수치나 통계",
            )

        with col_b:
            st.markdown("**4️⃣ 시놉시스 (선택)**")
            st.caption("비워두면 Claude가 자동 설계합니다.")
            for stage_key, stage_hint in stage_hooks:
                st.text_input(
                    stage_key,
                    placeholder=stage_hint,
                    key=f"synopsis_{stage_key}",
                )
            st.text_area(
                "5️⃣ 등장 캐릭터 (선택)",
                placeholder=(
                    "예: 주인공 외 대조 인물\n"
                    "- 영끌족 청년 (피해자)\n"
                    "- 갭투자자 (가해자)"
                ),
                height=80,
                key="planning_characters",
            )

    st.divider()

    col_in, col_out = st.columns([1, 2])

    # ── 입력 영역 ──────────────────────────────────────────────────────────────
    with col_in:
        st.subheader("⚙️ 입력 설정")

        benchmark_input = st.text_area(
            "벤치마킹 대상",
            placeholder=(
                "유튜브 영상 URL, 키워드, 또는 경쟁 채널명을 입력하세요.\n"
                "예) 부동산, https://youtube.com/..., 부자아빠"
            ),
            height=100,
            key="p1_benchmark_input",
        )

        video_length = st.radio(
            "원하는 영상 길이",
            ["10분 내외 (5,000자)", "20분 내외 (10,000자)", "30분 내외 (15,000자)"],
            index=1,
            key="p1_length",
        )

        extra_req = st.text_area(
            "추가 요구사항 (선택)",
            placeholder="예) 20~30대 타겟, 자극적 어조 자제, 데이터 중심",
            height=80,
            key="p1_extra",
        )

        run_btn = st.button(
            "🔍 주제 발굴 시작",
            type="primary",
            use_container_width=True,
            disabled=not (channel_name and benchmark_input),
        )

    # ── 결과 영역 ──────────────────────────────────────────────────────────────
    with col_out:
        st.subheader("📋 분석 결과")

        if run_btn:
            # ── 기획 정보 통합 ──
            synopsis_text = ""
            for stage_key, _ in stage_hooks:
                val = st.session_state.get(f"synopsis_{stage_key}", "")
                if val.strip():
                    synopsis_text += f"[{stage_key}] {val}\n"

            planning_context = ""
            if st.session_state.get("planning_topic"):
                planning_context += f"영상 주제: {st.session_state['planning_topic']}\n"
            if st.session_state.get("planning_message"):
                planning_context += f"핵심 메시지: {st.session_state['planning_message']}\n"
            if st.session_state.get("planning_emotion"):
                planning_context += f"타겟 감정 흐름: {st.session_state['planning_emotion']}\n"
            if synopsis_text:
                planning_context += f"\n시놉시스:\n{synopsis_text}\n"
            if st.session_state.get("planning_data"):
                planning_context += f"\n주요 데이터:\n{st.session_state['planning_data']}\n"
            if st.session_state.get("planning_characters"):
                planning_context += f"\n등장 캐릭터:\n{st.session_state['planning_characters']}\n"

            st.session_state["planning_context"] = planning_context

            with st.spinner("Claude AI가 주제를 분석하는 중... (10~30초 소요)"):
                try:
                    result = call_claude_prompt1(
                        channel_name, benchmark_input, video_length, extra_req,
                        planning_context,
                    )
                    st.session_state["p1_result"] = result
                    st.session_state[P1_CHANNEL]    = channel_name
                    st.session_state[P1_BENCHMARK]  = benchmark_input
                except json.JSONDecodeError as e:
                    st.error(f"JSON 파싱 오류: {e}\n다시 시도해주세요.")
                    st.stop()
                except Exception as e:
                    st.error(f"분석 오류: {e}")
                    st.stop()

        result = st.session_state.get("p1_result")
        if not result:
            st.info("좌측에서 채널과 벤치마킹 대상을 입력하고 분석을 시작하세요.")
        else:
            with st.expander("📊 주제 발굴 결과 보기 (클릭하여 펼치기)", expanded=False):
                # 최우선 탑픽 배너
                top_pick = result.get("top_pick", {})
                top_rank = top_pick.get("rank", 1)
                st.success(
                    f"🏆 **최우선 추천: 주제 {top_rank}**  —  {top_pick.get('reason','')}\n\n"
                    f"첫 24h 예상: **{top_pick.get('first_24h_views','')}**  |  "
                    f"7일 누적: **{top_pick.get('day7_views','')}**  |  "
                    f"구독 전환율: **{top_pick.get('subscribe_rate','')}**"
                )

                # 분석 방법론 패널
                render_methodology_panel(result.get("methodology", {}))

                # 주제 카드 — 탑픽 먼저
                topics = result.get("topics", [])
                top_topic = next((t for t in topics if t.get("rank") == top_rank), None)
                other_topics = [t for t in topics if t.get("rank") != top_rank]

                if top_topic:
                    render_topic_card(top_topic, is_top=True)
                for topic in sorted(other_topics, key=lambda x: x.get("rank", 99)):
                    render_topic_card(topic, is_top=False)

                # SEO 섹션
                seo = result.get("seo", {})
                with st.expander("📌 SEO 키워드 전략"):
                    st.markdown("**메인 키워드**")
                    st.write(", ".join(seo.get("main_keywords", [])))
                    st.markdown("**롱테일 키워드**")
                    for kw in seo.get("longtail_keywords", []):
                        st.write(f"• {kw}")
                    st.markdown("**해시태그 (10개)**")
                    st.write(" ".join(seo.get("hashtags", [])))

                # 내보내기
                st.divider()
                saved_channel = st.session_state.get(P1_CHANNEL, "")
                saved_benchmark = st.session_state.get(P1_BENCHMARK, "")
                fname = f"주제발굴_{saved_channel}_{datetime.now().strftime('%Y%m%d_%H%M')}"

                ec1, ec2 = st.columns(2)
                with ec1:
                    excel_bytes = export_to_excel(result, saved_channel, saved_benchmark)
                    st.download_button(
                        "📥 Excel 다운로드 (3시트)",
                        data=excel_bytes,
                        file_name=f"{fname}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                with ec2:
                    csv_lines = ["순위,제목,핵심메시지,검색량,경쟁도,CTR,적합도,난이도,Hook문장"]
                    for t in topics:
                        csv_lines.append(
                            f"{t.get('rank','')},{t.get('title','').replace(',','，')},"
                            f"{t.get('core_message','').replace(',','，')},"
                            f"{t.get('search_volume','')},{t.get('competition','')},"
                            f"{t.get('expected_ctr','')},{t.get('persona_fit','')},"
                            f"{t.get('difficulty','')},{t.get('hook_sentence','').replace(',','，')}"
                        )
                    st.download_button(
                        "📥 CSV 다운로드",
                        data="\n".join(csv_lines).encode("utf-8-sig"),
                        file_name=f"{fname}.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )

                with st.expander("📌 구글 드라이브 업로드 방법"):
                    st.markdown(
                        """
                        1. 위 버튼으로 파일 다운로드
                        2. [drive.google.com](https://drive.google.com) 접속
                        3. **+ 새로 만들기** → **파일 업로드** 클릭
                        4. 다운로드한 파일 선택 → 업로드 완료!

                        > **팁:** Excel 파일은 구글 드라이브에서 **구글 스프레드시트**로 바로 열 수 있습니다.
                        """
                    )

    # ── Stage 3: AI 분석 결과 → 기획 정보 자동 연결 ──────────────────────────
    if st.session_state.get("p1_result"):
        _r3 = st.session_state.get("p1_result", {})
        _top_pick3 = _r3.get("top_pick", {})
        _top_rank3 = _top_pick3.get("rank", 1)
        _topics3 = _r3.get("topics", [])
        _top_topic3 = next(
            (t for t in _topics3 if t.get("rank") == _top_rank3),
            _topics3[0] if _topics3 else {},
        )

        if _top_topic3:
            with st.expander("🔗 AI 분석 결과를 기획 정보에 자동 적용", expanded=False):
                st.write(
                    f"최우선 추천 주제 **'{_top_topic3.get('title', '')}'** 의 "
                    "분석 결과를 기획 정보 입력 필드에 자동으로 채웁니다."
                )
                _s3_col, _s3_btn_col = st.columns([3, 1])
                with _s3_col:
                    _apply_stage3 = st.checkbox(
                        "이 분석 결과를 기획 정보에 자동 적용",
                        value=True,
                        key="_stage3_checkbox",
                    )
                with _s3_btn_col:
                    _stage3_btn = st.button(
                        "📋 기획 정보에 적용 →",
                        use_container_width=True,
                        disabled=not _apply_stage3,
                        key="_stage3_apply_btn",
                    )
                if _stage3_btn and _apply_stage3:
                    st.session_state["planning_topic"] = _top_topic3.get("title", "")
                    st.session_state["planning_message"] = _top_topic3.get("core_message", "")
                    st.session_state["planning_emotion"] = _top_topic3.get("target_emotion", "")
                    st.session_state["_autofill_stage3_applied"] = True
                    st.session_state["_autofill_stage3_source"] = _top_topic3.get("title", "")
                    st.rerun()

    # ── 확정 버튼 (전체 너비, 컬럼 외부) ──────────────────────────────────────
    if st.session_state.get("p1_result"):
        st.divider()

        if not st.session_state.get(P1_TOPIC_TITLE):
            st.warning("⚠️ 위 결과에서 주제를 선택 후 확정하세요.")
        else:
            selected = st.session_state.get(P1_TOPIC_TITLE, "")
            st.success(f"✅ 선택된 주제: **{selected}**")

            # Stage 4: 확정 시 썸네일 탭 자동 연결 여부
            _apply_stage4 = st.checkbox(
                "확정 시 썸네일·제목 탭에 자동으로 연결",
                value=True,
                key="_stage4_checkbox",
            )

            if st.button(
                "✅ 이 주제로 확정하고 썸네일·제목 단계로 →",
                type="primary",
                use_container_width=True,
                key="confirm_topic",
            ):
                st.session_state["p1_confirmed"] = True
                if _apply_stage4:
                    st.session_state["_autofill_stage4_applied"] = True
                    st.session_state["_autofill_stage4_source"] = selected
                st.info("👆 상단에서 **🎨 썸네일·제목** 탭을 클릭하세요.")
                st.rerun()

        if st.session_state.get("p1_confirmed"):
            st.success("✅ 주제 확정 완료! **🎨 썸네일·제목** 탭으로 이동하세요.")
