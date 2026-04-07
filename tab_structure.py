import re
import json
import io
import os
from datetime import datetime

import streamlit as st
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from channel_db import CHANNEL_DB
from prompts import PROMPT_3_SYSTEM
from error_handler import handle_api_error
from session_state_manager import (
    P1_CHANNEL, P1_TOPIC_TITLE, P1_CORE_MESSAGE, P1_EMOTION, P1_HOOK,
    P2_TITLE, P2_THUMBNAIL, P2_HOOK_30SEC,
    P3_RESULT, P3_VIDEO_LENGTH, P3_STRUCTURE, P3_EMOTION_MAP, P3_MINI_HOOKS, P3_SCENE_META,
    render_pipeline_status, render_p1_confirmed_card, render_p2_confirmed_card,
)


# ──────────────────────────────────────────
# 상수
# ──────────────────────────────────────────

SECTION_CONFIG = {
    "hook":     {"label": "HOOK",     "timecode": "00:00", "color": "#E53935"},
    "teaser":   {"label": "TEASER",   "timecode": "01:00", "color": "#FB8C00"},
    "big_idea": {"label": "BIG IDEA", "timecode": "02:00", "color": "#F9A825"},
    "intro":    {"label": "INTRO",    "timecode": "03:00", "color": "#43A047"},
    "body1":    {"label": "BODY 1",   "timecode": "04:00", "color": "#00897B"},
    "body2":    {"label": "BODY 2",   "timecode": "07:00", "color": "#039BE5"},
    "body3":    {"label": "BODY 3",   "timecode": "10:15", "color": "#1E88E5"},
    "body4":    {"label": "BODY 4",   "timecode": "13:30", "color": "#5E35B1"},
    "reveal":   {"label": "REVEAL",   "timecode": "17:00", "color": "#8E24AA"},
    "impact":   {"label": "IMPACT",   "timecode": "18:30", "color": "#D81B60"},
    "end":      {"label": "END",      "timecode": "19:00", "color": "#546E7A"},
}

SECTION_ORDER = ["hook", "teaser", "big_idea", "intro", "body1", "body2", "body3", "body4", "reveal", "impact", "end"]
BODY_KEYS = ["body1", "body2", "body3", "body4"]
MINI_HOOK_MAP = {"07:00": "body2", "10:15": "body3", "13:30": "body4", "16:45": "reveal"}


# ──────────────────────────────────────────
# 클라이언트
# ──────────────────────────────────────────

def _get_client():
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        key = ""
    key = key or os.environ.get("ANTHROPIC_API_KEY", "")
    return anthropic.Anthropic(api_key=key)


# ──────────────────────────────────────────
# JSON 파싱 방어
# ──────────────────────────────────────────

def _extract_json(text: str) -> str:
    text = text.strip()
    m = re.search(r"```json\s*([\s\S]*?)\s*```", text)
    if m:
        return m.group(1).strip()
    m = re.search(r"```\s*([\s\S]*?)\s*```", text)
    if m:
        candidate = m.group(1).strip()
        if candidate.startswith("{"):
            return candidate
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return text[start:end + 1]
    return text


def _safe_loads(text: str) -> dict:
    json_str = _extract_json(text)
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        preview = json_str[:300].replace("\n", " ")
        raise json.JSONDecodeError(
            f"JSON 파싱 실패 (앞 300자: {preview}...)", e.doc, e.pos
        )


# ──────────────────────────────────────────
# 페르소나 블록 생성
# ──────────────────────────────────────────

def build_persona_block(channel_name: str) -> str:
    if channel_name not in CHANNEL_DB:
        return "채널 정보 없음 — 중립적 분석가 페르소나 적용"
    info = CHANNEL_DB[channel_name]
    return (
        f"채널명: {channel_name}\n"
        f"주인공: {info['host']} ({info['host_desc']})\n"
        f"색상: {info['color_primary']} / {info['color_secondary']}\n"
        f"톤앤매너: {info['tone']}\n"
        f"타겟: {info['target']}\n"
        f"썸네일 전략: {info['thumbnail_style']}\n"
        f"시각 무드: {info['visual_mood']}"
    )


# ──────────────────────────────────────────
# Claude API 호출 (3회 재시도)
# ──────────────────────────────────────────

def call_claude_prompt3(
    channel_name: str,
    topic_title: str,
    core_message: str,
    target_emotion: str,
    confirmed_title: str,
    confirmed_thumbnail: str,
    hook_30sec: str,
    video_length: str,
) -> dict:
    persona_block = build_persona_block(channel_name)
    system_prompt = PROMPT_3_SYSTEM.format(
        persona_block=persona_block,
        channel_name=channel_name,
        topic_title=topic_title,
        core_message=core_message,
        target_emotion=target_emotion,
        confirmed_title=confirmed_title,
        confirmed_thumbnail=confirmed_thumbnail,
        hook_30sec=hook_30sec,
        video_length=video_length,
    )

    user_message = (
        f"채널명: {channel_name}\n"
        f"확정 주제: {topic_title}\n"
        f"핵심 메시지: {core_message}\n"
        f"타겟 감정: {target_emotion}\n"
        f"확정 제목: {confirmed_title}\n"
        f"확정 썸네일 문구: {confirmed_thumbnail}\n"
        f"초반 30초 Hook: {hook_30sec}\n"
        f"원하는 영상 길이: {video_length}\n\n"
        "위 정보를 바탕으로 대본 구조를 설계하고 JSON만 반환하라.\n"
        "응답은 반드시 { 로 시작하고 } 로 끝나야 한다."
    )

    MAX_ATTEMPTS = 3
    last_raw = ""
    last_error = None
    client = _get_client()

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_message}],
            )
            last_raw = response.content[0].text
            return _safe_loads(last_raw)

        except json.JSONDecodeError as e:
            last_error = e
            if attempt < MAX_ATTEMPTS:
                user_message = (
                    f"이전 응답이 유효한 JSON이 아니었다.\n"
                    f"오류: {str(e)}\n"
                    f"이전 응답 앞부분: {last_raw[:300]}\n\n"
                    "규칙 재확인 후 올바른 JSON만 반환하라:\n"
                    "1. 응답 첫 글자는 반드시 { 이어야 한다\n"
                    "2. 마크다운 코드블록(```) 절대 사용 금지\n"
                    "3. 설명 텍스트 절대 금지\n\n"
                    f"채널명: {channel_name} / 주제: {topic_title}\n"
                    "JSON만 반환하라."
                )

        except Exception as e:
            raise RuntimeError(f"API 호출 오류: {str(e)}")

    raise ValueError(
        f"Claude API가 {MAX_ATTEMPTS}회 시도 후에도 유효한 JSON을 반환하지 못했습니다.\n"
        f"마지막 오류: {str(last_error)}\n"
        "해결 방법: 입력 내용을 더 간결하게 줄이거나 잠시 후 재시도하세요."
    )


# ──────────────────────────────────────────
# UI 컴포넌트: 섹션 카드
# ──────────────────────────────────────────

def render_section_card(key: str, sec: dict):
    cfg = SECTION_CONFIG.get(key, {})
    label = cfg.get("label", key.upper())
    timecode = cfg.get("timecode", "")
    color = cfg.get("color", "#888")
    word_count = sec.get("word_count", "")

    info_purpose = sec.get("info_purpose", "")
    emotion_goal = sec.get("emotion_goal", "")
    scene_type = sec.get("scene_type", "")
    protagonist_role = sec.get("protagonist_role", "")
    supporting = sec.get("supporting_characters", "")
    key_objects = sec.get("key_objects", "")
    retention = sec.get("retention_device", "")
    # Body 전용
    core_topic = sec.get("core_topic", "")
    consequence = sec.get("consequence", "")
    mini_hook_bridge = sec.get("mini_hook_bridge", "")
    # Reveal 전용
    strongest_truth = sec.get("strongest_truth", "")
    lingering_effect = sec.get("lingering_effect", "")
    # Impact 전용
    life_connection = sec.get("life_connection", "")
    key_reaction = sec.get("key_reaction", "")
    # End 전용
    final_message = sec.get("final_message", "")
    action_suggestion = sec.get("action_suggestion", "")
    emotion_close = sec.get("emotion_close", "")
    lingering_device = sec.get("lingering_device", "")

    def row(icon, label_text, value):
        if not value:
            return ""
        return (
            f'<div style="margin-bottom:5px;">'
            f'<span style="color:#888;font-size:11px;">{icon} {label_text}</span><br>'
            f'<span style="color:#1a1a1a;font-size:13px;">{value}</span>'
            f'</div>'
        )

    extra_rows = ""
    if core_topic:
        extra_rows += row("📌", "핵심 토픽", core_topic)
    if consequence:
        extra_rows += row("⚡", "결과·여파", consequence)
    if mini_hook_bridge:
        extra_rows += row("🪝", "미니훅 브릿지", mini_hook_bridge)
    if strongest_truth:
        extra_rows += row("💥", "가장 강한 진실", strongest_truth)
    if lingering_effect:
        extra_rows += row("🌊", "여운 효과", lingering_effect)
    if life_connection:
        extra_rows += row("🔗", "삶 연결", life_connection)
    if key_reaction:
        extra_rows += row("😮", "핵심 반응", key_reaction)
    if final_message:
        extra_rows += row("🎯", "최종 메시지", final_message)
    if action_suggestion:
        extra_rows += row("✅", "행동 제안", action_suggestion)
    if emotion_close:
        extra_rows += row("💫", "감정 마무리", emotion_close)
    if lingering_device:
        extra_rows += row("🎵", "여운 장치", lingering_device)

    st.markdown(
        f"""
        <div style="
            border-left: 5px solid {color};
            background: #ffffff;
            border-radius: 8px;
            padding: 14px 16px;
            margin-bottom: 10px;
            box-shadow: 0 1px 4px rgba(0,0,0,0.06);
            color: #1a1a1a;
        ">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:10px;">
                <span style="font-size:15px; font-weight:700; color:{color};">
                    [{timecode}] {label}
                </span>
                <span style="font-size:12px; color:#888; background:#f5f5f5;
                             padding:2px 10px; border-radius:12px;">
                    {word_count}자
                </span>
            </div>
            {row("🎯", "정보 목적", info_purpose)}
            {row("💭", "감정 목표", emotion_goal)}
            {row("🎬", "장면 유형", scene_type)}
            {row("🎭", "주인공 역할", protagonist_role)}
            {row("👥", "보조 인물", supporting)}
            {row("🔑", "핵심 오브젝트", key_objects)}
            {extra_rows}
            {row("🔒", "리텐션 장치", retention)}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_section_editor(key: str, sec: dict) -> dict:
    cfg = SECTION_CONFIG.get(key, {})
    label = cfg.get("label", key.upper())
    timecode = cfg.get("timecode", "")
    updated = dict(sec)

    with st.expander(f"✏️ [{timecode}] {label}", expanded=False):
        is_body = key in BODY_KEYS
        is_reveal = key == "reveal"
        is_impact = key == "impact"
        is_end = key == "end"

        if is_body:
            updated["core_topic"] = st.text_input(
                "핵심 토픽", value=sec.get("core_topic", ""), key=f"p3e_{key}_topic"
            )
        updated["info_purpose"] = st.text_area(
            "정보 목적", value=sec.get("info_purpose", ""), height=60, key=f"p3e_{key}_info"
        )
        updated["emotion_goal"] = st.text_input(
            "감정 목표", value=sec.get("emotion_goal", ""), key=f"p3e_{key}_emotion"
        )
        updated["scene_type"] = st.text_input(
            "장면 유형", value=sec.get("scene_type", ""), key=f"p3e_{key}_scene"
        )
        if not is_impact:
            updated["protagonist_role"] = st.text_input(
                "주인공 역할", value=sec.get("protagonist_role", ""), key=f"p3e_{key}_role"
            )
        if not is_impact:
            updated["supporting_characters"] = st.text_input(
                "보조 인물", value=sec.get("supporting_characters", ""), key=f"p3e_{key}_supporting"
            )
        if not is_impact and not is_end:
            updated["key_objects"] = st.text_input(
                "핵심 오브젝트", value=sec.get("key_objects", ""), key=f"p3e_{key}_objects"
            )
        if not is_impact:
            updated["retention_device"] = st.text_input(
                "리텐션 장치", value=sec.get("retention_device", ""), key=f"p3e_{key}_retention"
            )
        if is_body:
            updated["consequence"] = st.text_input(
                "결과·여파", value=sec.get("consequence", ""), key=f"p3e_{key}_consequence"
            )
            updated["mini_hook_bridge"] = st.text_input(
                "미니훅 브릿지", value=sec.get("mini_hook_bridge", ""), key=f"p3e_{key}_minihook"
            )
        if is_reveal:
            updated["strongest_truth"] = st.text_area(
                "가장 강한 진실", value=sec.get("strongest_truth", ""), height=60, key=f"p3e_{key}_truth"
            )
            updated["lingering_effect"] = st.text_input(
                "여운 효과", value=sec.get("lingering_effect", ""), key=f"p3e_{key}_linger"
            )
        if is_impact:
            updated["life_connection"] = st.text_area(
                "삶 연결", value=sec.get("life_connection", ""), height=60, key=f"p3e_{key}_life"
            )
            updated["key_reaction"] = st.text_input(
                "핵심 반응", value=sec.get("key_reaction", ""), key=f"p3e_{key}_reaction"
            )
        if is_end:
            updated["final_message"] = st.text_area(
                "최종 메시지", value=sec.get("final_message", ""), height=60, key=f"p3e_{key}_final"
            )
            updated["action_suggestion"] = st.text_input(
                "행동 제안", value=sec.get("action_suggestion", ""), key=f"p3e_{key}_action"
            )
            updated["emotion_close"] = st.text_input(
                "감정 마무리", value=sec.get("emotion_close", ""), key=f"p3e_{key}_eclose"
            )
            updated["lingering_device"] = st.text_input(
                "여운 장치", value=sec.get("lingering_device", ""), key=f"p3e_{key}_ldevice"
            )

    return updated


# ──────────────────────────────────────────
# UI 컴포넌트: 감정 지도
# ──────────────────────────────────────────

def render_emotion_map(emotion_map: list):
    if not emotion_map:
        return
    st.markdown("#### 🗺️ 감정 지도")
    cols_per_row = 3
    for i in range(0, len(emotion_map), cols_per_row):
        chunk = emotion_map[i:i + cols_per_row]
        cols = st.columns(len(chunk))
        for col, em in zip(cols, chunk):
            with col:
                emoji = em.get("emoji", "")
                emotion = em.get("emotion", "")
                timecode = em.get("timecode", "")
                desc = em.get("description", "")
                st.markdown(
                    f"""
                    <div style="
                        background:#f8f9fa; border-radius:8px;
                        padding:10px 12px; margin-bottom:8px;
                        border: 1px solid #e0e0e0; color:#1a1a1a;
                        text-align:center;
                    ">
                        <div style="font-size:24px;">{emoji}</div>
                        <div style="font-size:11px; color:#888; margin:2px 0;">[{timecode}]</div>
                        <div style="font-size:13px; font-weight:700; color:#333;">{emotion}</div>
                        <div style="font-size:11px; color:#555; margin-top:4px;">{desc}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )


# ──────────────────────────────────────────
# Excel 내보내기
# ──────────────────────────────────────────

def _header_style(cell, bg="4A90E2"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _data_style(cell, bg="FFFFFF"):
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.fill = PatternFill("solid", fgColor=bg)


def export_p3_excel(
    structure: dict,
    emotion_map: list,
    mini_hooks: list,
    topic_title: str,
    channel_name: str,
) -> bytes:
    wb = Workbook()

    # ── 시트 1: 구조 설계 ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📐 구조설계"

    info_rows = [
        ("채널명", channel_name),
        ("주제", topic_title),
        ("작성 일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (k, v) in enumerate(info_rows, 1):
        ws1.cell(i, 1, k).font = Font(bold=True)
        ws1.cell(i, 2, v)
    ws1.append([])

    headers = ["타임코드", "섹션", "글자수", "핵심 토픽", "정보 목적", "감정 목표",
               "장면 유형", "주인공 역할", "보조 인물", "핵심 오브젝트", "리텐션 장치", "결과·여파", "미니훅 브릿지"]
    row_start = len(info_rows) + 2
    ws1.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        _header_style(ws1.cell(row_start, col_idx))
    ws1.row_dimensions[row_start].height = 30

    for key in SECTION_ORDER:
        sec = structure.get(key, {})
        cfg = SECTION_CONFIG.get(key, {})
        bg = "F8F9FA" if SECTION_ORDER.index(key) % 2 == 0 else "FFFFFF"
        row_data = [
            cfg.get("timecode", ""),
            cfg.get("label", key.upper()),
            sec.get("word_count", ""),
            sec.get("core_topic", ""),
            sec.get("info_purpose", ""),
            sec.get("emotion_goal", ""),
            sec.get("scene_type", ""),
            sec.get("protagonist_role", ""),
            sec.get("supporting_characters", ""),
            sec.get("key_objects", ""),
            sec.get("retention_device", ""),
            sec.get("consequence", ""),
            sec.get("mini_hook_bridge", ""),
        ]
        ws1.append(row_data)
        for col_idx, _ in enumerate(row_data, 1):
            _data_style(ws1.cell(ws1.max_row, col_idx), bg)
        ws1.row_dimensions[ws1.max_row].height = 45

    col_widths = [10, 12, 8, 18, 28, 18, 20, 20, 20, 20, 20, 20, 22]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 2: 감정 지도 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("🗺️ 감정지도")
    em_headers = ["타임코드", "감정", "이모지", "설명"]
    ws2.append(em_headers)
    for col_idx, _ in enumerate(em_headers, 1):
        _header_style(ws2.cell(1, col_idx), bg="7C4DFF")
    ws2.row_dimensions[1].height = 30

    for idx, em in enumerate(emotion_map):
        bg = "F3E5F5" if idx % 2 == 0 else "FFFFFF"
        row_data = [
            em.get("timecode", ""),
            em.get("emotion", ""),
            em.get("emoji", ""),
            em.get("description", ""),
        ]
        ws2.append(row_data)
        for col_idx, _ in enumerate(row_data, 1):
            _data_style(ws2.cell(ws2.max_row, col_idx), bg)
        ws2.row_dimensions[ws2.max_row].height = 35

    for i, w in enumerate([12, 18, 8, 50], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 3: 미니훅 & 메타 ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("🪝 미니훅_메타")
    mh_headers = ["타임코드", "미니훅 문장", "장면 전환 기능", "다음 장면 유형"]
    ws3.append(mh_headers)
    for col_idx, _ in enumerate(mh_headers, 1):
        _header_style(ws3.cell(1, col_idx), bg="E53935")
    ws3.row_dimensions[1].height = 30

    for idx, mh in enumerate(mini_hooks):
        bg = "FFF3E0" if idx % 2 == 0 else "FFFFFF"
        row_data = [
            mh.get("timecode", ""),
            mh.get("sentence", ""),
            mh.get("scene_transition", ""),
            mh.get("next_scene_type", ""),
        ]
        ws3.append(row_data)
        for col_idx, _ in enumerate(row_data, 1):
            _data_style(ws3.cell(ws3.max_row, col_idx), bg)
        ws3.row_dimensions[ws3.max_row].height = 40

    for i, w in enumerate([12, 40, 30, 25], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────
# 메인 탭 렌더링
# ──────────────────────────────────────────

def render_structure_tab():
    render_pipeline_status()

    st.header("📐 대본 구조 설계", divider="gray")
    st.caption("썸네일·제목 약속을 Scene-First 방식으로 완벽하게 이행하는 대본 구조를 설계합니다.")

    # 이전 단계 확정 카드 표시
    ok1 = render_p1_confirmed_card(editable=False)
    if not ok1:
        return

    ok2 = render_p2_confirmed_card(editable=False)
    if not ok2:
        return

    st.divider()

    # ── 입력 설정 ─────────────────────────────────────────────────────────────
    with st.expander("⚙️ 대본 구조 설계 설정", expanded=True):
        col_left, col_right = st.columns([2, 1])
        with col_left:
            video_length = st.selectbox(
                "원하는 영상 길이",
                options=["10분 내외", "15분 내외", "20분 내외", "25분 내외", "30분 이상"],
                index=2,
                key="p3_video_length_select",
            )
        with col_right:
            st.markdown("&nbsp;", unsafe_allow_html=True)

        extra_notes = st.text_area(
            "추가 요청사항 (선택)",
            placeholder="예: 초반 Hook을 더 강하게, BODY 2에서 통계 데이터 강조 등",
            height=70,
            key="p3_extra_notes",
        )

    run_btn = st.button(
        "🚀 대본 구조 설계 시작",
        key="p3_run_btn",
        type="primary",
        use_container_width=True,
    )

    if run_btn:
        channel_name = st.session_state.get(P1_CHANNEL, "")
        topic_title = st.session_state.get(P1_TOPIC_TITLE, "")
        core_message = st.session_state.get(P1_CORE_MESSAGE, "")
        target_emotion = st.session_state.get(P1_EMOTION, "")
        confirmed_title = st.session_state.get(P2_TITLE, "")
        confirmed_thumbnail = st.session_state.get(P2_THUMBNAIL, "")
        hook_30sec = st.session_state.get(P2_HOOK_30SEC, "")

        with st.spinner("Claude AI가 대본 구조를 설계 중입니다... (20-40초 소요)"):
            try:
                result = call_claude_prompt3(
                    channel_name=channel_name,
                    topic_title=topic_title,
                    core_message=core_message,
                    target_emotion=target_emotion,
                    confirmed_title=confirmed_title,
                    confirmed_thumbnail=confirmed_thumbnail,
                    hook_30sec=hook_30sec,
                    video_length=video_length,
                )
                st.session_state[P3_RESULT] = result
                st.session_state[P3_VIDEO_LENGTH] = video_length
                st.session_state[P3_STRUCTURE] = result.get("structure", {})
                st.session_state[P3_EMOTION_MAP] = result.get("emotion_map", [])
                st.session_state[P3_MINI_HOOKS] = result.get("mini_hooks", [])
                st.session_state[P3_SCENE_META] = result.get("scene_meta", {})
                st.success("대본 구조 설계 완료!")
            except ValueError as e:
                st.error(f"분석 오류: {e}")
            except RuntimeError as e:
                handle_api_error(e, context="대본 구조 설계")
            except Exception as e:
                handle_api_error(e, context="대본 구조 설계")

    result = st.session_state.get(P3_RESULT)
    if not result:
        st.info("왼쪽 설정을 확인한 후 '대본 구조 설계 시작' 버튼을 눌러주세요.", icon="👆")
        return

    structure = st.session_state.get(P3_STRUCTURE, {})
    emotion_map = st.session_state.get(P3_EMOTION_MAP, [])
    mini_hooks = st.session_state.get(P3_MINI_HOOKS, [])
    scene_meta = st.session_state.get(P3_SCENE_META, {})
    channel_name = st.session_state.get(P1_CHANNEL, "")
    topic_title = st.session_state.get(P1_TOPIC_TITLE, "")
    video_length = st.session_state.get(P3_VIDEO_LENGTH, "20분 내외")

    # ── 섹션 1: 썸네일 약속 분석 ──────────────────────────────────────────────
    st.subheader("📋 썸네일 약속 분석")
    thumb_promise = result.get("thumbnail_promise", {})
    if thumb_promise:
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown(
                f"""
                <div style="background:#fff3cd; border-radius:8px; padding:14px;
                            border-left:4px solid #ffc107; color:#1a1a1a;">
                    <div style="font-weight:700; margin-bottom:6px;">🎯 핵심 약속</div>
                    <div style="font-size:14px;">{thumb_promise.get('core_promise','')}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown("&nbsp;", unsafe_allow_html=True)
            st.markdown(f"**🔘 보조 약속 1:** {thumb_promise.get('sub_promise_1','')}")
            st.markdown(f"**🔘 보조 약속 2:** {thumb_promise.get('sub_promise_2','')}")
        with col_b:
            st.markdown(f"**👆 클릭 이유:** {thumb_promise.get('click_reason','')}")
            st.markdown(f"**⏰ 회수 타이밍:** {thumb_promise.get('recovery_timing','')}")
            if thumb_promise.get("scene_recovery"):
                st.info(f"🎬 장면 회수: {thumb_promise.get('scene_recovery','')}")
            if thumb_promise.get("explain_recovery"):
                st.info(f"💬 설명 회수: {thumb_promise.get('explain_recovery','')}")

    # ── 섹션 2: 초반 30초 Hook 설계 ────────────────────────────────────────────
    st.subheader("⚡ 초반 30초 Hook 설계")
    hook_design = result.get("hook_design", {})
    if hook_design:
        phase_tabs = st.tabs(["Phase 1 (0~20초)", "Phase 2 (20~40초)", "Phase 3 (40~60초)"])
        phase_keys = ["phase1_0_20sec", "phase2_20_40sec", "phase3_40_60sec"]
        phase_labels = [
            ["core_scene", "core_sentence", "keywords", "scene_type", "protagonist_role"],
            ["shock_method", "evidence", "scene_type", "supporting_character"],
            ["bigger_question", "tension_point", "scene_type", "key_object"],
        ]
        label_names = {
            "core_scene": "핵심 회수 장면", "core_sentence": "핵심 문장",
            "keywords": "필수 키워드", "scene_type": "장면 유형",
            "protagonist_role": "주인공 역할", "shock_method": "충격 입증 방식",
            "evidence": "사용할 증거", "supporting_character": "보조 인물/반응",
            "bigger_question": "더 큰 질문", "tension_point": "긴장 포인트",
            "key_object": "핵심 오브젝트",
        }
        for tab, pk, fields in zip(phase_tabs, phase_keys, phase_labels):
            with tab:
                phase_data = hook_design.get(pk, {})
                for field in fields:
                    val = phase_data.get(field, "")
                    if val:
                        st.markdown(f"**{label_names.get(field, field)}:** {val}")

    # ── 섹션 3: 8단계 대본 구조 ────────────────────────────────────────────────
    st.subheader("🗂️ 8단계 대본 구조")

    view_mode = st.radio(
        "보기 모드",
        options=["카드 뷰 (읽기 전용)", "편집 모드"],
        horizontal=True,
        key="p3_view_mode",
    )

    if view_mode == "카드 뷰 (읽기 전용)":
        cols = st.columns(2)
        for i, key in enumerate(SECTION_ORDER):
            sec = structure.get(key, {})
            with cols[i % 2]:
                render_section_card(key, sec)
    else:
        updated_structure = {}
        for key in SECTION_ORDER:
            sec = structure.get(key, {})
            updated_structure[key] = render_section_editor(key, sec)

        if st.button("✅ 구조 수정 확정", key="p3_confirm_edit", type="primary"):
            st.session_state[P3_STRUCTURE] = updated_structure
            st.success("수정된 구조가 저장되었습니다!")

    # ── 섹션 4: 감정 지도 ──────────────────────────────────────────────────────
    st.divider()
    render_emotion_map(emotion_map)

    emotion_analysis = result.get("emotion_map_analysis", "")
    if emotion_analysis:
        with st.expander("📊 감정 지도 분석", expanded=False):
            st.write(emotion_analysis)

    # ── 섹션 5: 미니훅 배치 ────────────────────────────────────────────────────
    if mini_hooks:
        st.divider()
        st.subheader("🪝 미니훅 배치")
        mini_cols = st.columns(len(mini_hooks))
        for col, mh in zip(mini_cols, mini_hooks):
            with col:
                st.markdown(
                    f"""
                    <div style="
                        background:#f0f7ff; border-radius:8px;
                        padding:12px; border: 1px solid #4A90E2;
                        color: #1a1a1a;
                    ">
                        <div style="font-size:11px; color:#4A90E2; font-weight:700;
                                    margin-bottom:6px;">[{mh.get('timecode','')}]</div>
                        <div style="font-size:13px; font-weight:600; margin-bottom:8px;">
                            "{mh.get('sentence','')}"
                        </div>
                        <div style="font-size:11px; color:#666;">
                            🔄 {mh.get('scene_transition','')}<br>
                            ▶ 다음: {mh.get('next_scene_type','')}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    # ── 섹션 6: Scene Meta (프롬프트 4/5 연동) ────────────────────────────────
    if scene_meta:
        st.divider()
        with st.expander("🎬 Scene Meta — 프롬프트 4·5 연동 정보", expanded=False):
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                roles = scene_meta.get("protagonist_roles", {})
                if roles:
                    st.markdown("**주인공 역할 분류**")
                    for rk, rv in roles.items():
                        if rv:
                            st.caption(f"• {rk}: {rv}")
                supporting_cast = scene_meta.get("supporting_cast", [])
                if supporting_cast:
                    st.markdown("**보조 인물 설계**")
                    for cast in supporting_cast:
                        st.caption(
                            f"• {cast.get('name','')} — {cast.get('role','')} [{cast.get('emotion','')}]"
                        )
            with col_m2:
                key_objects = scene_meta.get("key_visual_objects", [])
                if key_objects:
                    st.markdown("**핵심 시각 오브젝트**")
                    for obj in key_objects:
                        st.caption(f"• {obj}")
                for field, label in [
                    ("consequence_sections", "결과·여파 강조 구간"),
                    ("crowd_reaction_sections", "군중·사회 반응 구간"),
                    ("evidence_sections", "사물·증거 중심 구간"),
                ]:
                    val = scene_meta.get(field, "")
                    if val:
                        st.markdown(f"**{label}:** {val}")

            p4_inst = scene_meta.get("prompt4_instruction", "")
            p5_inst = scene_meta.get("prompt5_instruction", "")
            if p4_inst:
                st.info(f"📝 대본 작성(P4) 지시: {p4_inst}")
            if p5_inst:
                st.info(f"🎥 영상 기획(P5) 지시: {p5_inst}")

    # ── 확정 및 내보내기 ────────────────────────────────────────────────────────
    st.divider()
    st.subheader("✅ 확정 및 내보내기")

    confirm_col, export_col = st.columns(2)
    with confirm_col:
        if st.button("✅ 이 구조로 확정하고 다음 단계로", key="p3_confirm_final", type="primary"):
            st.session_state[P3_STRUCTURE] = structure
            st.success("대본 구조가 확정되었습니다! 다음 탭(대본 작성)으로 이동하세요.")
            st.balloons()

    with export_col:
        try:
            excel_bytes = export_p3_excel(
                structure=structure,
                emotion_map=emotion_map,
                mini_hooks=mini_hooks,
                topic_title=topic_title,
                channel_name=channel_name,
            )
            fname = f"대본구조_{topic_title[:20]}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="📥 Excel 내보내기",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="p3_download_excel",
            )
        except Exception as e:
            st.warning(f"Excel 생성 오류: {e}")
