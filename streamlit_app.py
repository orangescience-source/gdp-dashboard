import io
import os
import json
from datetime import datetime
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tab_topic import render_topic_tab
from tab_thumbnail import render_thumbnail_tab
from tab_structure import render_structure_tab

st.set_page_config(
    page_title="YouTube 니치 발굴 대시보드",
    page_icon="🎬",
    layout="wide",
)

# ── API Key 처리 ──────────────────────────────────────────────────────────────

def get_api_key() -> str:
    """st.secrets → 환경변수 → 사이드바 입력 순으로 API 키를 가져온다."""
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
        if key:
            return key
    except Exception:
        pass
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if key:
        return key
    return ""

# ── 프롬프트 & Claude 호출 ────────────────────────────────────────────────────

SYSTEM_PROMPT = """당신은 유튜브 채널 전략 전문가입니다.
사용자가 제공하는 키워드를 기반으로 유망한 유튜브 니치 주제를 분석하고,
반드시 아래 JSON 형식으로만 응답하세요. 다른 텍스트는 절대 포함하지 마세요."""

JSON_SCHEMA = """{
  "niches": [
    {
      "name": "니치명 (한국어)",
      "description": "이 니치에 대한 2-3문장 설명",
      "competition": 경쟁도_숫자(1-10),
      "monetization": 수익성_숫자(1-10),
      "trend": 트렌드_숫자(1-10),
      "opportunity_score": 종합기회점수_숫자(1-10),
      "estimated_monthly_views": "예상 월 조회수 범위 (예: 5만~20만)",
      "content_ideas": ["아이디어1", "아이디어2", "아이디어3", "아이디어4", "아이디어5"],
      "target_audience": "주요 시청자층 설명",
      "pros": ["장점1", "장점2", "장점3"],
      "cons": ["단점1", "단점2"],
      "recommended_format": "쇼츠 / 롱폼 / 혼합",
      "posting_frequency": "권장 업로드 빈도 (예: 주 2-3회)"
    }
  ],
  "market_summary": "전반적인 시장 분석 요약 (3-4문장)",
  "top_recommendation": "가장 추천하는 니치명"
}"""

def build_prompt(keywords: str, n: int) -> str:
    return f"""사용자 관심 키워드: {keywords}
분석할 니치 수: {n}개

위 키워드를 바탕으로 유튜브 채널 개설에 적합한 니치 주제 {n}개를 분석해주세요.

각 항목 평가 기준:
- competition (경쟁도): 1=매우 낮음(유리), 10=매우 높음(불리)
- monetization (수익성): 1=낮음, 10=매우 높음
- trend (트렌드): 1=하락 중, 10=급성장 중
- opportunity_score (종합 기회 점수): 경쟁도 낮고, 수익성 높고, 트렌드 높을수록 높은 점수

반드시 아래 JSON 형식으로만 응답하세요:
{JSON_SCHEMA}"""


def analyze_niches(api_key: str, keywords: str, n: int) -> dict:
    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": build_prompt(keywords, n)}],
    )
    raw = response.content[0].text.strip()
    # JSON 블록이 코드 펜스로 감싸진 경우 제거
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())

# ── 차트 헬퍼 ─────────────────────────────────────────────────────────────────

def bubble_chart(df: pd.DataFrame):
    fig = px.scatter(
        df,
        x="competition",
        y="monetization",
        size="trend",
        color="opportunity_score",
        text="name",
        color_continuous_scale="RdYlGn",
        size_max=50,
        labels={
            "competition": "경쟁도 (낮을수록 유리)",
            "monetization": "수익성 (높을수록 유리)",
            "trend": "트렌드",
            "opportunity_score": "기회 점수",
        },
        title="니치 포지셔닝 맵 — 경쟁도 vs 수익성",
    )
    fig.update_traces(textposition="top center")
    fig.update_layout(height=460, coloraxis_colorbar=dict(title="기회 점수"))
    return fig


def opportunity_bar(df: pd.DataFrame):
    df_sorted = df.sort_values("opportunity_score", ascending=True)
    fig = px.bar(
        df_sorted,
        x="opportunity_score",
        y="name",
        orientation="h",
        color="opportunity_score",
        color_continuous_scale="Blues",
        text="opportunity_score",
        labels={"opportunity_score": "종합 기회 점수", "name": ""},
        title="니치별 종합 기회 점수",
    )
    fig.update_traces(texttemplate="%{text:.1f}", textposition="outside")
    fig.update_layout(height=380, showlegend=False, coloraxis_showscale=False)
    return fig


def radar_chart(niche: dict):
    categories = ["수익성", "트렌드", "기회점수", "경쟁 낮음"]
    values = [
        niche["monetization"],
        niche["trend"],
        niche["opportunity_score"],
        10 - niche["competition"] + 1,  # 경쟁도 반전
    ]
    fig = go.Figure(
        go.Scatterpolar(
            r=values + [values[0]],
            theta=categories + [categories[0]],
            fill="toself",
            fillcolor="rgba(99,110,250,0.3)",
            line_color="rgb(99,110,250)",
        )
    )
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 10])),
        showlegend=False,
        height=280,
        margin=dict(l=40, r=40, t=40, b=40),
    )
    return fig

# ── Excel 내보내기 ────────────────────────────────────────────────────────────

def style_header_cell(cell, bg_color="1F3864", font_color="FFFFFF"):
    cell.font = Font(bold=True, color=font_color, size=11)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data_cell(cell, bg_color="FFFFFF"):
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_excel(result: dict, keywords: str) -> bytes:
    niches = result.get("niches", [])
    top_rec = result.get("top_recommendation", "")
    market_summary = result.get("market_summary", "")

    wb = Workbook()

    # ── 시트 1: 요약 ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 분석 요약"

    ws1.merge_cells("A1:F1")
    ws1["A1"] = "🎬 YouTube 니치 발굴 분석 결과"
    ws1["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1["A2"] = "분석 키워드"
    ws1["B2"] = keywords
    ws1["A3"] = "분석 일시"
    ws1["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws1["A4"] = "최추천 니치"
    ws1["B4"] = top_rec
    ws1["A5"] = "시장 요약"
    ws1["B5"] = market_summary
    ws1["B5"].alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[5].height = 60

    for row in ws1["A2:B5"]:
        for cell in row:
            thin = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws1["A2:A5"]:
        cell[0].font = Font(bold=True)
        cell[0].fill = PatternFill("solid", fgColor="E8F0FE")

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 60

    ws1.append([])
    headers_sum = ["순위", "니치명", "경쟁도\n(낮을수록 유리)", "수익성", "트렌드", "종합 기회 점수", "예상 월 조회수", "추천 포맷"]
    ws1.append(headers_sum)
    for i, h in enumerate(headers_sum, 1):
        style_header_cell(ws1.cell(ws1.max_row, i))
    ws1.row_dimensions[ws1.max_row].height = 30

    sorted_niches = sorted(niches, key=lambda x: x["opportunity_score"], reverse=True)
    for rank, niche in enumerate(sorted_niches, 1):
        row = [
            rank,
            niche["name"],
            niche["competition"],
            niche["monetization"],
            niche["trend"],
            niche["opportunity_score"],
            niche.get("estimated_monthly_views", ""),
            niche.get("recommended_format", ""),
        ]
        ws1.append(row)
        bg = "FFF9C4" if niche["name"] == top_rec else ("F8F9FA" if rank % 2 == 0 else "FFFFFF")
        for col_idx, _ in enumerate(row, 1):
            style_data_cell(ws1.cell(ws1.max_row, col_idx), bg)

    for col_idx, width in enumerate([6, 22, 12, 10, 10, 14, 18, 12], 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 시트 2: 상세 분석 ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("🗂️ 상세 분석")
    detail_headers = [
        "니치명", "설명", "시청자층",
        "경쟁도", "수익성", "트렌드", "종합 기회 점수",
        "예상 월 조회수", "추천 포맷", "업로드 빈도",
        "장점", "단점",
    ]
    ws2.append(detail_headers)
    for i, _ in enumerate(detail_headers, 1):
        style_header_cell(ws2.cell(1, i))
    ws2.row_dimensions[1].height = 30

    for idx, niche in enumerate(sorted_niches):
        row = [
            niche["name"],
            niche.get("description", ""),
            niche.get("target_audience", ""),
            niche["competition"],
            niche["monetization"],
            niche["trend"],
            niche["opportunity_score"],
            niche.get("estimated_monthly_views", ""),
            niche.get("recommended_format", ""),
            niche.get("posting_frequency", ""),
            " / ".join(niche.get("pros", [])),
            " / ".join(niche.get("cons", [])),
        ]
        ws2.append(row)
        bg = "F8F9FA" if idx % 2 == 0 else "FFFFFF"
        for col_idx, _ in enumerate(row, 1):
            style_data_cell(ws2.cell(ws2.max_row, col_idx), bg)
        ws2.row_dimensions[ws2.max_row].height = 50

    for col_idx, width in enumerate([22, 45, 30, 10, 10, 10, 14, 18, 12, 14, 35, 25], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 시트 3: 콘텐츠 아이디어 ───────────────────────────────────────────────
    ws3 = wb.create_sheet("💡 콘텐츠 아이디어")
    ws3.append(["니치명", "콘텐츠 아이디어 1", "아이디어 2", "아이디어 3", "아이디어 4", "아이디어 5"])
    for i in range(1, 7):
        style_header_cell(ws3.cell(1, i))
    ws3.row_dimensions[1].height = 30

    for idx, niche in enumerate(sorted_niches):
        ideas = niche.get("content_ideas", [])
        ideas += [""] * (5 - len(ideas))
        row = [niche["name"]] + ideas[:5]
        ws3.append(row)
        bg = "F8F9FA" if idx % 2 == 0 else "FFFFFF"
        for col_idx, _ in enumerate(row, 1):
            style_data_cell(ws3.cell(ws3.max_row, col_idx), bg)
        ws3.row_dimensions[ws3.max_row].height = 40

    for col_idx, width in enumerate([22, 30, 30, 30, 30, 30], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_csv(result: dict, keywords: str) -> bytes:
    niches = result.get("niches", [])
    rows = []
    for niche in sorted(niches, key=lambda x: x["opportunity_score"], reverse=True):
        rows.append({
            "키워드": keywords,
            "니치명": niche["name"],
            "설명": niche.get("description", ""),
            "경쟁도": niche["competition"],
            "수익성": niche["monetization"],
            "트렌드": niche["trend"],
            "종합 기회 점수": niche["opportunity_score"],
            "예상 월 조회수": niche.get("estimated_monthly_views", ""),
            "시청자층": niche.get("target_audience", ""),
            "추천 포맷": niche.get("recommended_format", ""),
            "업로드 빈도": niche.get("posting_frequency", ""),
            "장점": " / ".join(niche.get("pros", [])),
            "단점": " / ".join(niche.get("cons", [])),
            "콘텐츠 아이디어": " | ".join(niche.get("content_ideas", [])),
            "최추천 여부": "✅" if niche["name"] == result.get("top_recommendation") else "",
        })
    return pd.DataFrame(rows).to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


# ── 페이지 레이아웃 ───────────────────────────────────────────────────────────

# 전역 CSS: 탭 글자 크기 확대 + 커스텀 카드 가독성
st.markdown("""
<style>
/* 탭 버튼 글자 크기·굵기 */
.stTabs [data-baseweb="tab"] {
    font-size: 16px !important;
    font-weight: 600 !important;
    padding: 10px 22px !important;
}
.stTabs [data-baseweb="tab"]:hover {
    color: #4A90E2 !important;
}
/* 커스텀 HTML 카드 내 기본 텍스트 강제 */
.custom-card, .custom-card * {
    box-sizing: border-box;
}
</style>
""", unsafe_allow_html=True)

st.title("🎬 YouTube 채널 전략 도구")
st.caption("Claude AI가 분석하는 유튜브 채널 전략 플랫폼")

tab1, tab2, tab3, tab4 = st.tabs(["🔍 니치 발굴", "📊 주제 발굴", "🎨 썸네일·제목", "📐 대본 구조"])

# ── 사이드바 (니치 발굴 탭용) ─────────────────────────────────────────────────

with st.sidebar:
    st.header("⚙️ 분석 설정")

    api_key = get_api_key()
    if not api_key:
        api_key = st.text_input(
            "Anthropic API Key",
            type="password",
            placeholder="sk-ant-...",
            help="Anthropic 콘솔에서 발급한 API 키를 입력하세요.",
        )

    st.divider()

    keywords = st.text_area(
        "관심 키워드",
        placeholder="예: 요리, 1인 가구, 간단 레시피\n예: 재테크, 주식, 20대\n예: 헬스, 다이어트, 홈트",
        height=120,
        help="분석하고 싶은 관심사나 주제를 자유롭게 입력하세요.",
    )

    n_niches = st.slider("발굴할 니치 수", min_value=3, max_value=8, value=5)

    run_btn = st.button("🔍 니치 발굴 시작", use_container_width=True, type="primary")

    st.divider()
    st.caption(
        "점수 기준\n"
        "- 경쟁도: 낮을수록 진입 유리 (1-10)\n"
        "- 수익성: 높을수록 수익화 쉬움 (1-10)\n"
        "- 트렌드: 높을수록 성장 중 (1-10)\n"
        "- 기회 점수: 세 지표 종합 (1-10)"
    )

# ── 탭 1: 니치 발굴 ──────────────────────────────────────────────────────────

with tab1:
    if run_btn:
        if not api_key:
            st.error("API 키를 입력해주세요.")
        elif not keywords.strip():
            st.warning("관심 키워드를 입력해주세요.")
        else:
            with st.spinner("Claude AI가 니치를 분석 중입니다... (10-30초 소요)"):
                try:
                    result = analyze_niches(api_key, keywords.strip(), n_niches)
                    st.session_state["result"] = result
                except json.JSONDecodeError as e:
                    st.error(f"응답 파싱 오류: {e}\n다시 시도해주세요.")
                except Exception as e:
                    st.error(f"분석 오류: {e}")

    if "result" in st.session_state:
        result = st.session_state["result"]
        niches = result.get("niches", [])
        top_rec = result.get("top_recommendation", "")
        market_summary = result.get("market_summary", "")

        if not niches:
            st.warning("분석 결과가 없습니다. 다시 시도해주세요.")
            st.stop()

        df = pd.DataFrame(niches)

        # ── 섹션 1: 요약 메트릭 ──────────────────────────────────────────────
        st.header("📊 분석 요약", divider="gray")

        col1, col2, col3, col4 = st.columns(4)
        avg_comp = df["competition"].mean()
        avg_mono = df["monetization"].mean()
        avg_trend = df["trend"].mean()

        col1.metric("🏆 최추천 니치", top_rec)
        col2.metric("📊 평균 경쟁도", f"{avg_comp:.1f} / 10", help="낮을수록 진입 유리")
        col3.metric("💰 평균 수익성", f"{avg_mono:.1f} / 10")
        col4.metric("📈 평균 트렌드", f"{avg_trend:.1f} / 10")

        if market_summary:
            with st.expander("📝 시장 분석 요약 보기"):
                st.write(market_summary)

        st.divider()

        # ── 섹션 2: 차트 ─────────────────────────────────────────────────────
        st.header("📈 시각화 분석", divider="gray")

        chart_col1, chart_col2 = st.columns(2)
        with chart_col1:
            st.plotly_chart(bubble_chart(df), use_container_width=True)
        with chart_col2:
            st.plotly_chart(opportunity_bar(df), use_container_width=True)

        st.divider()

        # ── 섹션 3: 니치 카드 ────────────────────────────────────────────────
        st.header("🗂️ 니치 상세 분석", divider="gray")

        for niche in sorted(niches, key=lambda x: x["opportunity_score"], reverse=True):
            badge = "⭐ 최추천" if niche["name"] == top_rec else ""
            with st.expander(f"**{niche['name']}** {badge}  —  기회 점수: {niche['opportunity_score']}/10"):
                left, right = st.columns([2, 1])

                with left:
                    st.write(niche["description"])
                    st.write(f"**시청자층:** {niche['target_audience']}")
                    st.write(f"**예상 월 조회수:** {niche['estimated_monthly_views']}")
                    st.write(f"**추천 포맷:** {niche['recommended_format']}  |  **업로드 빈도:** {niche['posting_frequency']}")
                    st.write("**점수**")
                    score_data = {
                        "경쟁도 (낮을수록 유리)": niche["competition"],
                        "수익성": niche["monetization"],
                        "트렌드": niche["trend"],
                        "종합 기회 점수": niche["opportunity_score"],
                    }
                    for label, val in score_data.items():
                        st.progress(int(val * 10), text=f"{label}: {val}/10")

                with right:
                    st.plotly_chart(radar_chart(niche), use_container_width=True)

                idea_col, pro_col, con_col = st.columns(3)
                with idea_col:
                    st.write("**💡 콘텐츠 아이디어**")
                    for idea in niche.get("content_ideas", []):
                        st.write(f"- {idea}")
                with pro_col:
                    st.write("**✅ 장점**")
                    for pro in niche.get("pros", []):
                        st.write(f"- {pro}")
                with con_col:
                    st.write("**⚠️ 단점**")
                    for con in niche.get("cons", []):
                        st.write(f"- {con}")

        # ── 섹션 4: 내보내기 ─────────────────────────────────────────────────
        st.header("☁️ 구글 드라이브로 내보내기", divider="gray")
        st.caption("파일을 다운로드한 후 구글 드라이브에 업로드하세요.")

        fname_base = f"youtube_niche_{datetime.now().strftime('%Y%m%d_%H%M')}"

        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            excel_bytes = generate_excel(result, keywords)
            st.download_button(
                label="📥 Excel로 다운로드 (.xlsx)",
                data=excel_bytes,
                file_name=f"{fname_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
            st.caption("3개 시트: 요약 / 상세 분석 / 콘텐츠 아이디어")

        with dl_col2:
            csv_bytes = generate_csv(result, keywords)
            st.download_button(
                label="📥 CSV로 다운로드 (.csv)",
                data=csv_bytes,
                file_name=f"{fname_base}.csv",
                mime="text/csv",
                use_container_width=True,
            )
            st.caption("엑셀/구글 시트에서 바로 열 수 있는 형식")

        with st.expander("📌 구글 드라이브 업로드 방법"):
            st.markdown(
                """
                1. 위 버튼으로 파일 다운로드
                2. [drive.google.com](https://drive.google.com) 접속
                3. **+ 새로 만들기** → **파일 업로드** 클릭
                4. 다운로드한 파일 선택 → 업로드 완료!

                > **팁:** Excel 파일은 구글 드라이브에서 **구글 스프레드시트**로 바로 열 수 있습니다.
                """
            )

    else:
        st.info(
            "왼쪽 사이드바에서 API 키와 관심 키워드를 입력한 후 **'니치 발굴 시작'** 버튼을 눌러주세요.",
            icon="👈",
        )
        st.markdown(
            """
            ### 이 앱으로 할 수 있는 것
            | 기능 | 설명 |
            |------|------|
            | 🔍 키워드 기반 니치 발굴 | 관심사 키워드로 유망 유튜브 니치 발견 |
            | 📊 경쟁도 분석 | 각 니치의 진입 장벽과 경쟁 강도 평가 |
            | 💰 수익성 분석 | 광고 수익 및 수익화 가능성 평가 |
            | 📈 트렌드 분석 | 현재 성장 중인 니치 파악 |
            | 🗂️ 콘텐츠 아이디어 | 각 니치별 구체적인 영상 아이디어 제공 |
            | ☁️ 구글 드라이브 내보내기 | Excel/CSV 다운로드 후 드라이브에 저장 |
            """
        )

# ── 탭 2: 주제 발굴 ──────────────────────────────────────────────────────────

with tab2:
    render_topic_tab()

# ── 탭 3: 썸네일·제목 전략 ────────────────────────────────────────────────────

with tab3:
    render_thumbnail_tab()


# ── 탭 4: 대본 구조 설계 ──────────────────────────────────────────────────────

with tab4:
    render_structure_tab()
