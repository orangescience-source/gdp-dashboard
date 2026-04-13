import io
import json
import os
from datetime import datetime

import streamlit as st
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from channel_db import CHANNEL_DB
from prompts import PROMPT_6_SYSTEM
from session_state_manager import (
    P1_CHANNEL, P1_TOPIC_TITLE, P1_CORE_MESSAGE, P1_EMOTION, P1_HOOK,
    P2_RESULT, P2_TITLE, P2_THUMBNAIL, P2_HOOK_30SEC,
    P3_STRUCTURE, P3_EMOTION_MAP, P3_MINI_HOOKS, P3_SCENE_META,
    P4_SCRIPT_FULL, P4_VIZ_MEMO, P4_CONFIRMED,
    P6_RESULT, P6_FINAL_TITLE, P6_DESCRIPTION, P6_HASHTAGS, P6_CONFIRMED,
    render_pipeline_status,
)


# ──────────────────────────────────────────
# API 클라이언트
# ──────────────────────────────────────────

def _get_client():
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        key = ""
    key = key or os.environ.get("ANTHROPIC_API_KEY", "")
    return anthropic.Anthropic(api_key=key)


# ──────────────────────────────────────────
# JSON 파싱 3단계 방어 로직
# ──────────────────────────────────────────

def _extract_json(raw: str) -> dict:
    text = raw.strip()
    # 1단계: 마크다운 코드펜스 제거
    if text.startswith("```"):
        parts = text.split("```")
        for part in parts:
            part = part.strip()
            if part.startswith("json"):
                part = part[4:].strip()
            try:
                return json.loads(part)
            except json.JSONDecodeError:
                continue
    # 2단계: 전체 직접 파싱
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    # 3단계: 중괄호 균형 탐색
    start = text.find("{")
    if start == -1:
        raise json.JSONDecodeError("No JSON found", text, 0)
    depth, end, in_str, esc = 0, -1, False, False
    for i, ch in enumerate(text[start:], start):
        if esc:
            esc = False
            continue
        if ch == "\\":
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i
                break
    if end == -1:
        raise json.JSONDecodeError("Unbalanced braces", text, start)
    return json.loads(text[start:end + 1])


# ──────────────────────────────────────────
# 채널 페르소나 & SEO 헬퍼
# ──────────────────────────────────────────

def _build_persona_block(channel_name: str) -> str:
    if channel_name not in CHANNEL_DB:
        return "채널 정보 없음 — 중립 SEO 전략 적용"
    info = CHANNEL_DB[channel_name]
    return (
        f"채널명: {channel_name}\n"
        f"톤앤매너: {info['tone']}\n"
        f"타겟: {info['target']}\n"
        f"커뮤니티 톤: {info.get('community_tone', '')}\n"
        f"커뮤니티 퀴즈 예시: {info.get('community_quiz', '')}"
    )


def _get_seo_keywords(channel_name: str) -> str:
    if channel_name not in CHANNEL_DB:
        return ""
    return ", ".join(CHANNEL_DB[channel_name].get("seo_keywords", []))


def _structure_summary(structure: list) -> str:
    return "\n".join(
        f"[{s.get('timestamp_start','')}] STAGE {s.get('stage','')} "
        f"{s.get('section','')} — {s.get('title','')}"
        for s in structure
    )


# ──────────────────────────────────────────
# 완성도 체크리스트
# ──────────────────────────────────────────

def _render_checklist() -> int:
    checks = [
        ("주제 발굴 완료",       bool(st.session_state.get(P1_TOPIC_TITLE)),   "탭2(주제 발굴)에서 주제를 확정해주세요."),
        ("채널 선택 완료",       bool(st.session_state.get(P1_CHANNEL)),       "탭2(주제 발굴)에서 채널을 선택해주세요."),
        ("핵심 메시지 확정",     bool(st.session_state.get(P1_CORE_MESSAGE)),  "탭2(주제 발굴)에서 핵심 메시지를 확정해주세요."),
        ("썸네일 문구 확정",     bool(st.session_state.get(P2_THUMBNAIL)),     "탭3(썸네일·제목)에서 썸네일 문구를 확정해주세요."),
        ("영상 제목 확정",       bool(st.session_state.get(P2_TITLE)),         "탭3(썸네일·제목)에서 제목을 확정해주세요."),
        ("대본 구조 확정",       bool(st.session_state.get(P3_STRUCTURE)),     "탭4(대본 구조)에서 구조를 확정해주세요."),
        ("감정 지도 확정",       bool(st.session_state.get(P3_EMOTION_MAP)),   "탭4(대본 구조)에서 감정 지도를 확정해주세요."),
        ("미니훅 확정",          bool(st.session_state.get(P3_MINI_HOOKS)),    "탭4(대본 구조)에서 미니훅을 확정해주세요."),
        ("전체 대본 생성 완료",  bool(st.session_state.get(P4_SCRIPT_FULL)),   "탭5(대본 작성)에서 대본을 생성해주세요."),
        ("전체 대본 확정 저장",  bool(st.session_state.get(P4_CONFIRMED)),     "탭5(대본 작성)에서 '전체 대본 확정 저장'을 눌러주세요."),
    ]

    done_count = sum(1 for _, v, _ in checks if v)
    pct = int(done_count / len(checks) * 100)
    bar_color = "#2ecc71" if pct == 100 else ("#f39c12" if pct >= 60 else "#e74c3c")

    st.markdown(
        f"""
        <div style="margin-bottom:12px;">
            <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
                <span style="font-size:13px; font-weight:600; color:#333;">완성도 체크리스트</span>
                <span style="font-size:13px; font-weight:700; color:{bar_color};">
                    {done_count}/{len(checks)} ({pct}%)
                </span>
            </div>
            <div style="background:#e9ecef; border-radius:6px; height:10px;">
                <div style="width:{pct}%; background:{bar_color};
                            border-radius:6px; height:10px;"></div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    for i, (label, done, hint) in enumerate(checks):
        with (col1 if i % 2 == 0 else col2):
            icon = "✅" if done else "⬜"
            color = "#155724" if done else "#856404"
            bg    = "#d4edda" if done else "#fff3cd"
            hint_html = f'<br><span style="font-size:11px;color:#777;">{hint}</span>' if not done else ""
            st.markdown(
                f'<div style="background:{bg};color:{color};padding:6px 10px;'
                f'border-radius:6px;margin-bottom:6px;font-size:13px;">'
                f'{icon} {label}{hint_html}</div>',
                unsafe_allow_html=True,
            )
    return done_count


# ──────────────────────────────────────────
# P1~P4 요약 테이블
# ──────────────────────────────────────────

def _render_summary_table():
    structure  = st.session_state.get(P3_STRUCTURE, [])
    script     = st.session_state.get(P4_SCRIPT_FULL, "")

    rows = [
        ("채널",           st.session_state.get(P1_CHANNEL, "-")),
        ("확정 주제",      st.session_state.get(P1_TOPIC_TITLE, "-")),
        ("핵심 메시지",    st.session_state.get(P1_CORE_MESSAGE, "-")),
        ("확정 제목",      st.session_state.get(P2_TITLE, "-")),
        ("썸네일 문구",    (st.session_state.get(P2_THUMBNAIL, "") or "-").replace("\n", " / ")),
        ("초반 30초 Hook", st.session_state.get(P2_HOOK_30SEC, "-")),
        ("대본 구조",      f"{len(structure)}단계"),
        ("대본 분량",      f"{len(script):,}자" if script else "-"),
    ]

    html_rows = "".join(
        f'<tr><td style="font-weight:600;color:#555;padding:7px 12px;'
        f'border-bottom:1px solid #eee;white-space:nowrap;">{k}</td>'
        f'<td style="padding:7px 12px;border-bottom:1px solid #eee;color:#222;">{v}</td></tr>'
        for k, v in rows
    )
    st.markdown(
        f'<div style="border:1px solid #ddd;border-radius:10px;overflow:hidden;margin-bottom:16px;">'
        f'<table style="width:100%;border-collapse:collapse;font-size:13px;background:#fff;">'
        f'{html_rows}</table></div>',
        unsafe_allow_html=True,
    )


# ──────────────────────────────────────────
# Claude API 호출 (3회 재시도)
# ──────────────────────────────────────────

def call_claude_prompt6(
    channel_name, topic_title, core_message,
    video_title, thumbnail_text, structure,
) -> dict:
    system_prompt = PROMPT_6_SYSTEM.format(
        persona_block     = _build_persona_block(channel_name),
        seo_keywords      = _get_seo_keywords(channel_name),
        channel_name      = channel_name,
        topic_title       = topic_title,
        core_message      = core_message,
        video_title       = video_title,
        thumbnail_text    = thumbnail_text,
        structure_summary = _structure_summary(structure),
    )
    user_message = (
        f"채널: {channel_name} / 주제: {topic_title} / 제목: {video_title}\n"
        "위 정보로 유튜브 업로드 패키지를 생성하라. JSON만 반환하라."
    )

    MAX_ATTEMPTS = 3
    last_raw, last_error = "", None
    client = _get_client()

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=4096,
                system=system_prompt,
                messages=[{"role": "user", "content": user_message}],
            )
            last_raw = resp.content[0].text
            return _extract_json(last_raw)
        except json.JSONDecodeError as e:
            last_error = e
            if attempt < MAX_ATTEMPTS:
                user_message = (
                    f"이전 응답이 유효한 JSON이 아니었다.\n오류: {e}\n"
                    f"앞부분: {last_raw[:300]}\n\n"
                    "규칙: 응답은 { 로 시작하고 } 로 끝나야 한다. JSON만 반환하라."
                )
        except Exception as e:
            raise RuntimeError(f"API 호출 오류: {e}")

    raise ValueError(
        f"Claude API가 {MAX_ATTEMPTS}회 시도 후에도 유효한 JSON을 반환하지 못했습니다.\n"
        f"마지막 오류: {last_error}"
    )


# ──────────────────────────────────────────
# 결과 표시 컴포넌트
# ──────────────────────────────────────────

def _render_result(result: dict):
    st.subheader("🏷️ SEO 최적화 최종 제목")
    final_title = st.text_input(
        "최종 제목 (수정 가능)",
        value=result.get("final_title", ""),
        key="p6_edit_title",
    )

    st.subheader("📄 유튜브 설명란")
    desc_val = result.get("description", "").replace("\\n", "\n")
    description = st.text_area(
        "설명란 (수정 가능 — 타임라인 포함)",
        value=desc_val,
        height=300,
        key="p6_edit_desc",
    )

    st.subheader("🔖 해시태그 10개")
    hashtags = result.get("hashtags", [])
    edited_hashtags = st.text_input(
        "해시태그 (수정 가능, 스페이스 구분)",
        value=" ".join(hashtags),
        key="p6_edit_hashtags",
    )
    st.caption(f"총 {len(hashtags)}개")

    st.divider()

    st.subheader("📢 커뮤니티 게시물")
    col_prev, col_quiz = st.columns(2)
    with col_prev:
        st.markdown("**예고 게시물**")
        st.text_area(
            "",
            value=result.get("community_preview", "").replace("\\n", "\n"),
            height=150,
            key="p6_community_preview",
            label_visibility="collapsed",
        )
    with col_quiz:
        st.markdown("**참여 유도 퀴즈**")
        st.text_area(
            "",
            value=result.get("community_quiz", "").replace("\\n", "\n"),
            height=150,
            key="p6_community_quiz",
            label_visibility="collapsed",
        )

    st.divider()

    st.subheader("🛍️ 추천 제품 태그 (타임스탬프 분산)")
    products = result.get("products", [])
    if products:
        cols = st.columns(len(products))
        for i, prod in enumerate(products):
            with cols[i]:
                st.markdown(
                    f'<div style="border:1px solid #ddd;border-radius:10px;padding:14px;'
                    f'background:#f8f9fa;color:#1a1a1a;text-align:center;">'
                    f'<div style="font-size:20px;margin-bottom:6px;">🛒</div>'
                    f'<div style="font-size:13px;font-weight:700;color:#111;margin-bottom:6px;">'
                    f'{prod.get("name","")}</div>'
                    f'<div style="font-size:12px;color:#e65100;font-weight:600;margin-bottom:6px;">'
                    f'⏱️ {prod.get("timestamp","")}</div>'
                    f'<div style="font-size:11px;color:#555;">{prod.get("reason","")}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    return final_title, description, edited_hashtags


# ──────────────────────────────────────────
# Excel 7시트 내보내기
# ──────────────────────────────────────────

def _hdr(cell, bg="2C3E50"):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _dat(cell, wrap=True):
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _row(ws, row_idx, values, header=False):
    for col, val in enumerate(values, 1):
        c = ws.cell(row_idx, col, str(val) if val is not None else "")
        _hdr(c) if header else _dat(c)


def generate_excel() -> bytes:
    wb = openpyxl.Workbook()
    structure   = st.session_state.get(P3_STRUCTURE, [])
    emotion_map = st.session_state.get(P3_EMOTION_MAP, [])
    mini_hooks  = st.session_state.get(P3_MINI_HOOKS, [])
    scene_meta  = st.session_state.get(P3_SCENE_META, [])
    full_script = st.session_state.get(P4_SCRIPT_FULL, "")
    viz_memo    = st.session_state.get(P4_VIZ_MEMO, "")
    p2_result   = st.session_state.get(P2_RESULT, {})

    # 시트1: 기획 개요
    ws1 = wb.active
    ws1.title = "기획개요"
    _row(ws1, 1, ["항목", "내용"], header=True)
    overview = [
        ("채널명",          st.session_state.get(P1_CHANNEL, "")),
        ("확정 주제",       st.session_state.get(P1_TOPIC_TITLE, "")),
        ("핵심 메시지",     st.session_state.get(P1_CORE_MESSAGE, "")),
        ("타겟 감정",       st.session_state.get(P1_EMOTION, "")),
        ("Hook 문장",       st.session_state.get(P1_HOOK, "")),
        ("확정 제목",       st.session_state.get(P2_TITLE, "")),
        ("썸네일 문구",     st.session_state.get(P2_THUMBNAIL, "")),
        ("초반 30초 Hook",  st.session_state.get(P2_HOOK_30SEC, "")),
        ("SEO 최종 제목",   st.session_state.get(P6_FINAL_TITLE, "")),
        ("해시태그",        " ".join(st.session_state.get(P6_HASHTAGS, []))),
        ("유튜브 설명란",   st.session_state.get(P6_DESCRIPTION, "")),
    ]
    for i, (k, v) in enumerate(overview, 2):
        _row(ws1, i, [k, v])
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 70

    # 시트2: 대본 구조
    ws2 = wb.create_sheet("대본구조")
    _row(ws2, 1, ["STAGE", "시작", "종료", "섹션", "제목", "목적", "내용가이드", "감정", "강도"], header=True)
    for i, s in enumerate(structure, 2):
        _row(ws2, i, [
            s.get("stage"), s.get("timestamp_start"), s.get("timestamp_end"),
            s.get("section"), s.get("title"), s.get("purpose"),
            s.get("content_guide"), s.get("emotion_target"), s.get("emotion_intensity"),
        ])
    for col, w in zip("ABCDEFGHI", [8, 8, 8, 14, 28, 35, 45, 12, 8]):
        ws2.column_dimensions[col].width = w

    # 시트3: 감정 지도
    ws3 = wb.create_sheet("감정지도")
    _row(ws3, 1, ["타임코드", "감정", "강도", "트리거", "STAGE"], header=True)
    for i, e in enumerate(emotion_map, 2):
        _row(ws3, i, [
            e.get("timestamp"), e.get("emotion"), e.get("intensity"),
            e.get("trigger"), e.get("stage"),
        ])
    for col, w in zip("ABCDE", [10, 12, 8, 40, 8]):
        ws3.column_dimensions[col].width = w

    # 시트4: 미니훅
    ws4 = wb.create_sheet("미니훅")
    _row(ws4, 1, ["타임코드", "유형", "훅 문장", "목적", "STAGE"], header=True)
    for i, h in enumerate(mini_hooks, 2):
        _row(ws4, i, [
            h.get("timestamp"), h.get("type"), h.get("hook_line"),
            h.get("purpose"), h.get("stage"),
        ])
    for col, w in zip("ABCDE", [10, 14, 50, 35, 8]):
        ws4.column_dimensions[col].width = w

    # 시트5: 완성 대본
    ws5 = wb.create_sheet("완성대본")
    ws5.column_dimensions["A"].width = 120
    _row(ws5, 1, ["전체 대본"], header=True)
    for i, line in enumerate(full_script.split("\n"), 2):
        c = ws5.cell(i, 1, line)
        _dat(c, wrap=False)

    # 시트6: 시각화 메모
    ws6 = wb.create_sheet("시각화메모")
    ws6.column_dimensions["A"].width = 120
    _row(ws6, 1, ["시각화 연동 메모"], header=True)
    memo_lines = viz_memo.split("\n") if viz_memo else ["(시각화 메모 없음)"]
    for i, line in enumerate(memo_lines, 2):
        c = ws6.cell(i, 1, line)
        _dat(c, wrap=False)

    # 시트7: 이미지 프롬프트
    ws7 = wb.create_sheet("이미지프롬프트")
    image_prompts = p2_result.get("image_prompts", []) if isinstance(p2_result, dict) else []
    if image_prompts:
        _row(ws7, 1, ["ID", "썸네일ID", "컨셉", "말풍선", "1행", "2행", "Full Prompt EN"], header=True)
        for i, ip in enumerate(image_prompts, 2):
            to = ip.get("text_overlay", {})
            _row(ws7, i, [
                ip.get("id"), ip.get("thumbnail_id"), ip.get("concept"),
                to.get("speech_bubble"), to.get("line1"), to.get("line2"),
                ip.get("full_prompt_en"),
            ])
        for ci, w in enumerate([6, 10, 20, 20, 20, 20, 80], 1):
            ws7.column_dimensions[get_column_letter(ci)].width = w
    else:
        ws7.cell(1, 1, "이미지 프롬프트 데이터 없음 (탭3 썸네일·제목 생성 후 채워집니다)")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────
# TXT 단일 파일 내보내기
# ──────────────────────────────────────────

def generate_txt() -> str:
    channel  = st.session_state.get(P1_CHANNEL, "")
    topic    = st.session_state.get(P1_TOPIC_TITLE, "")
    title    = st.session_state.get(P2_TITLE, "")
    thumbnail = st.session_state.get(P2_THUMBNAIL, "")
    script   = st.session_state.get(P4_SCRIPT_FULL, "")
    viz_memo = st.session_state.get(P4_VIZ_MEMO, "")
    final_t  = st.session_state.get(P6_FINAL_TITLE, "")
    desc     = st.session_state.get(P6_DESCRIPTION, "").replace("\\n", "\n")
    hashtags = " ".join(st.session_state.get(P6_HASHTAGS, []))
    now      = datetime.now().strftime("%Y-%m-%d %H:%M")

    sep = "=" * 50
    parts = [
        sep, f"  YouTube 기획 패키지 — {channel}", f"  생성일시: {now}", sep, "",
        "▶ 확정 주제", topic, "",
        "▶ 확정 제목 (원본)", title, "",
        "▶ SEO 최적화 제목", final_t, "",
        "▶ 썸네일 문구", thumbnail, "",
        "▶ 해시태그", hashtags, "",
        sep, "  유튜브 설명란", sep,
        desc, "",
        sep, "  전체 대본", sep,
        script, "",
    ]
    if viz_memo:
        parts += [sep, "  시각화 연동 메모", sep, viz_memo]
    return "\n".join(parts)


# ──────────────────────────────────────────
# 메인 탭 렌더링
# ──────────────────────────────────────────

def render_upload_tab():
    if not st.session_state.get("p4_confirmed"):
        st.info("📝 탭5에서 대본을 먼저 완성하고 확정해주세요.")
        st.stop()

    render_pipeline_status()

    st.header("📦 업로드 패키지")
    st.caption(
        "P1~P4 전체 기획 내용을 바탕으로 SEO 최적화 제목·설명란·해시태그·커뮤니티 게시물을 생성하고 "
        "Excel 7시트 + TXT 파일로 내보냅니다."
    )

    # ── 완성도 체크리스트 ──
    with st.expander("📋 완성도 체크리스트", expanded=True):
        _render_checklist()

    # ── 요약 테이블 ──
    with st.expander("📊 기획 내용 요약", expanded=False):
        _render_summary_table()

    st.divider()

    # ── 생성 버튼 ──
    channel_name = st.session_state.get(P1_CHANNEL, "")
    topic_title  = st.session_state.get(P1_TOPIC_TITLE, "")
    core_message = st.session_state.get(P1_CORE_MESSAGE, "")
    video_title  = st.session_state.get(P2_TITLE, "")
    thumbnail    = st.session_state.get(P2_THUMBNAIL, "")
    structure    = st.session_state.get(P3_STRUCTURE, [])

    can_generate = bool(channel_name and topic_title and video_title)

    run_btn = st.button(
        "📦 업로드 패키지 생성",
        type="primary",
        use_container_width=True,
        disabled=not can_generate,
        key="btn_upload_gen",
    )

    if run_btn:
        with st.spinner("Claude AI가 업로드 패키지를 생성하는 중... (10~20초 소요)"):
            try:
                result = call_claude_prompt6(
                    channel_name, topic_title, core_message,
                    video_title, thumbnail, structure,
                )
                st.session_state[P6_RESULT] = result
                st.success("✅ 업로드 패키지 생성 완료!")
            except ValueError as e:
                st.error(str(e))
                return
            except RuntimeError as e:
                st.error(f"API 오류: {e}")
                return
            except Exception as e:
                st.error(f"예기치 않은 오류: {e}")
                return

    result = st.session_state.get(P6_RESULT)
    if not result:
        st.info("위 버튼을 눌러 업로드 패키지를 생성하세요.")
        return

    st.divider()

    # ── 결과 표시 ──
    final_title, description, edited_hashtags = _render_result(result)

    st.divider()

    # ── 확정 저장 ──
    st.subheader("✅ 업로드 패키지 확정")
    if st.session_state.get(P6_CONFIRMED):
        st.success("✅ 이미 확정된 업로드 패키지가 있습니다. 재확정하려면 아래 버튼을 누르세요.")

    if st.button(
        "✅ 업로드 패키지 확정 저장",
        type="primary",
        use_container_width=True,
        key="confirm_upload",
    ):
        hashtag_list = [t.strip() for t in edited_hashtags.split() if t.strip()]
        st.session_state[P6_FINAL_TITLE] = final_title
        st.session_state[P6_DESCRIPTION] = description.replace("\\n", "\n")
        st.session_state[P6_HASHTAGS]    = hashtag_list
        st.session_state[P6_CONFIRMED]   = True
        st.success(f"✅ 업로드 패키지 확정! 해시태그 {len(hashtag_list)}개")
        st.balloons()

    st.divider()

    # ── 내보내기 ──
    st.subheader("📥 내보내기")
    topic_safe = (st.session_state.get(P1_TOPIC_TITLE, "upload") or "upload")[:20]
    now_str = datetime.now().strftime("%Y%m%d_%H%M")

    dl_col1, dl_col2 = st.columns(2)

    with dl_col1:
        try:
            excel_bytes = generate_excel()
            st.download_button(
                label="📊 Excel 다운로드 (7시트)",
                data=excel_bytes,
                file_name=f"youtube_package_{topic_safe}_{now_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                key="dl_excel",
            )
            st.caption("기획개요 / 대본구조 / 감정지도 / 미니훅 / 완성대본 / 시각화메모 / 이미지프롬프트")
        except Exception as e:
            st.error(f"Excel 생성 오류: {e}")

    with dl_col2:
        try:
            txt_content = generate_txt()
            st.download_button(
                label="📄 TXT 다운로드 (전체 통합)",
                data=txt_content.encode("utf-8"),
                file_name=f"youtube_package_{topic_safe}_{now_str}.txt",
                mime="text/plain",
                use_container_width=True,
                key="dl_txt",
            )
            st.caption("기획 개요 + 전체 대본 + 시각화 메모 단일 파일")
        except Exception as e:
            st.error(f"TXT 생성 오류: {e}")
