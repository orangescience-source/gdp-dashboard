import re
import json
import io
import os
import streamlit as st
import anthropic
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

from channel_db import CHANNEL_DB
from prompts import PROMPT_2_SYSTEM
from session_state_manager import (
    P1_CHANNEL, P1_TOPIC_TITLE, P1_CORE_MESSAGE, P1_EMOTION, P1_HOOK,
    P2_RESULT, P2_THUMBNAIL, P2_TITLE, P2_HOOK_30SEC, P2_IMAGE_PROMPT,
    render_pipeline_status, render_p1_confirmed_card,
)


def _get_client():
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        key = ""
    key = key or os.environ.get("ANTHROPIC_API_KEY", "")
    return anthropic.Anthropic(api_key=key)


# ──────────────────────────────────────────
# JSON 파싱 다층 방어
# ──────────────────────────────────────────

def _extract_json(text: str) -> str:
    text = text.strip()
    m = re.search(r"```json\s*([\s\S]*?)\s*```", text)
    if m:
        return m.group(1).strip()
    m = re.search(r"```\s*([\s\S]*?)\s*```", text)
    if m:
        candidate = m.group(1).strip()
        if candidate.startswith("{"):
            return candidate
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return text[start:end + 1]
    return text


def _safe_loads(text: str) -> dict:
    json_str = _extract_json(text)
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        preview = json_str[:200].replace("\n", " ")
        raise json.JSONDecodeError(
            f"JSON 파싱 실패 (앞 200자: {preview}...)", e.doc, e.pos
        )


def build_persona_block(channel_name: str) -> str:
    if channel_name not in CHANNEL_DB:
        return "채널 정보 없음 — 중립적 분석가 페르소나 적용"
    info = CHANNEL_DB[channel_name]
    return (
        f"채널명: {channel_name}\n"
        f"주인공: {info['host']} ({info['host_desc']})\n"
        f"색상: {info['color_primary']} / {info['color_secondary']}\n"
        f"톤앤매너: {info['tone']}\n"
        f"타겟: {info['target']}\n"
        f"썸네일 전략: {info['thumbnail_style']}\n"
        f"시각 무드: {info['visual_mood']}"
    )


# ──────────────────────────────────────────
# Claude API 호출 (3회 재시도 방어)
# ──────────────────────────────────────────

def call_claude_prompt2(channel_name, topic_title, core_message,
                        target_emotion, hook_sentence, extra_req=""):
    persona_block = build_persona_block(channel_name)
    system_prompt = PROMPT_2_SYSTEM.format(
        persona_block=persona_block,
        topic_title=topic_title,
        core_message=core_message,
        target_emotion=target_emotion,
        hook_sentence=hook_sentence,
    )

    user_message = (
        f"채널명: {channel_name}\n"
        f"주제: {topic_title}\n"
        f"핵심 메시지: {core_message}\n"
        f"타겟 감정: {target_emotion}\n"
        f"추가 요구사항: {extra_req if extra_req else '없음'}\n\n"
        "위 정보로 썸네일·제목 전략을 분석하고 JSON만 반환하라.\n"
        "응답은 반드시 { 로 시작하고 } 로 끝나야 한다."
    )

    MAX_ATTEMPTS = 3
    last_raw = ""
    last_error = None
    client = _get_client()

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_message}],
            )
            last_raw = response.content[0].text
            return _safe_loads(last_raw)

        except json.JSONDecodeError as e:
            last_error = e
            if attempt < MAX_ATTEMPTS:
                user_message = (
                    f"이전 응답이 유효한 JSON이 아니었다.\n"
                    f"오류: {str(e)}\n"
                    f"이전 응답 앞부분: {last_raw[:300]}\n\n"
                    "규칙 재확인 후 올바른 JSON만 반환하라:\n"
                    "1. 응답 첫 글자는 반드시 { 이어야 한다\n"
                    "2. 마크다운 코드블록(```) 절대 사용 금지\n"
                    "3. 설명 텍스트 절대 금지\n\n"
                    f"채널명: {channel_name} / 주제: {topic_title}\n"
                    "JSON만 반환하라."
                )

        except Exception as e:
            raise RuntimeError(f"API 호출 오류: {str(e)}")

    raise ValueError(
        f"Claude API가 {MAX_ATTEMPTS}회 시도 후에도 유효한 JSON을 반환하지 못했습니다.\n"
        f"마지막 오류: {str(last_error)}\n"
        "해결 방법: 입력 내용을 더 간결하게 줄이거나 잠시 후 재시도하세요."
    )


# ──────────────────────────────────────────
# UI 컴포넌트: 썸네일 문구 카드
# ──────────────────────────────────────────

def render_thumbnail_card(thumb: dict, selected_id: int):
    tid = thumb.get("id", 0)
    is_selected = (tid == selected_id)
    border = "2px solid #4A90E2" if is_selected else "1px solid #e0e0e0"
    bg = "#f0f7ff" if is_selected else "#ffffff"
    selected_badge = "✅ 선택됨" if is_selected else ""

    color_map = {
        "노란색": "#FFD700", "흰색": "#FFFFFF",
        "연두색": "#90EE90", "빨간색": "#FF4444",
    }

    def color_dot(color_name):
        hex_c = color_map.get(color_name, "#888")
        border_s = "border:1px solid #ccc;" if color_name == "흰색" else ""
        return (
            f'<span style="display:inline-block; width:12px; height:12px;'
            f' border-radius:50%; background:{hex_c}; {border_s}'
            f' margin-right:4px; vertical-align:middle;"></span>'
        )

    st.markdown(
        f"""
        <div style="
            border:{border}; border-radius:10px;
            padding:14px; margin-bottom:10px;
            background:{bg};
        ">
            <div style="display:flex; justify-content:space-between; margin-bottom:8px;">
                <span style="font-size:12px; color:#888;">문구 {tid} — {thumb.get('type','')}</span>
                <span style="font-size:12px; color:#4A90E2; font-weight:600;">{selected_badge}</span>
            </div>
            <div style="font-size:13px; margin-bottom:6px;">
                💬 말풍선: <b>{thumb.get('speech_bubble','')}</b>
                &nbsp; {color_dot(thumb.get('speech_bubble_color',''))}
                <span style="font-size:11px; color:#888;">{thumb.get('speech_bubble_color','')}</span>
            </div>
            <div style="font-size:13px; margin-bottom:4px;">
                📝 1행: <b>{thumb.get('line1','')}</b>
                &nbsp; {color_dot(thumb.get('line1_color',''))}
                <span style="font-size:11px; color:#888;">{thumb.get('line1_color','')}</span>
            </div>
            <div style="font-size:13px; margin-bottom:8px;">
                📝 2행: <b>{thumb.get('line2','')}</b>
                &nbsp; {color_dot(thumb.get('line2_color',''))}
                <span style="font-size:11px; color:#888;">{thumb.get('line2_color','')}</span>
            </div>
            <div style="display:flex; gap:8px;">
                <span style="font-size:11px; background:#eef; padding:3px 8px; border-radius:20px;">
                    총 {thumb.get('total_chars',0)}자
                </span>
                <span style="font-size:11px; background:#efe; padding:3px 8px; border-radius:20px;">
                    CTR {thumb.get('expected_ctr','')}
                </span>
                <span style="font-size:11px; background:#fff3cd; padding:3px 8px; border-radius:20px;">
                    장면적합 {thumb.get('scene_fit','')}
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ──────────────────────────────────────────
# Excel 내보내기
# ──────────────────────────────────────────

def export_p2_excel(result: dict, channel_name: str, topic_title: str) -> bytes:
    wb = openpyxl.Workbook()
    hf = PatternFill("solid", fgColor="2C3E50")
    hfont = Font(color="FFFFFF", bold=True, size=11)

    # 시트1: 썸네일 문구
    ws1 = wb.active
    ws1.title = "썸네일 문구"
    for col, h in enumerate(
        ["ID", "유형", "말풍선", "말풍선색", "1행", "1행색", "2행", "2행색", "글자수", "CTR", "장면적합"], 1
    ):
        c = ws1.cell(1, col, h)
        c.fill = hf
        c.font = hfont
    for i, t in enumerate(result.get("thumbnails", []), 2):
        row = [
            t.get("id"), t.get("type"), t.get("speech_bubble"), t.get("speech_bubble_color"),
            t.get("line1"), t.get("line1_color"), t.get("line2"), t.get("line2_color"),
            t.get("total_chars"), t.get("expected_ctr"), t.get("scene_fit"),
        ]
        for col, val in enumerate(row, 1):
            ws1.cell(i, col, str(val) if val is not None else "")

    # 시트2: 제목
    ws2 = wb.create_sheet("제목")
    for col, h in enumerate(["ID", "제목", "메인키워드", "감정장치", "검색적합도"], 1):
        c = ws2.cell(1, col, h)
        c.fill = hf
        c.font = hfont
    for i, t in enumerate(result.get("titles", []), 2):
        row = [t.get("id"), t.get("title"), t.get("main_keyword"),
               t.get("emotion_device"), t.get("search_fit")]
        for col, val in enumerate(row, 1):
            ws2.cell(i, col, str(val) if val is not None else "")

    # 시트3: Best 조합 + Hook
    ws3 = wb.create_sheet("추천조합_Hook")
    ws3.cell(1, 1, "채널명").font = Font(bold=True)
    ws3.cell(1, 2, channel_name)
    ws3.cell(2, 1, "주제").font = Font(bold=True)
    ws3.cell(2, 2, topic_title)
    for col, h in enumerate(["추천조합", "썸네일ID", "제목ID", "선정이유"], 1):
        c = ws3.cell(4, col, h)
        c.fill = hf
        c.font = hfont
    for i, combo in enumerate(result.get("best_combinations", []), 5):
        ws3.cell(i, 1, combo.get("rank"))
        ws3.cell(i, 2, combo.get("thumbnail_id"))
        ws3.cell(i, 3, combo.get("title_id"))
        ws3.cell(i, 4, combo.get("reason", ""))
    hook = result.get("hook_30sec", {})
    c = ws3.cell(10, 1, "초반Hook")
    c.fill = hf
    c.font = hfont
    ws3.cell(11, 1, "첫문장")
    ws3.cell(11, 2, hook.get("first_sentence", ""))
    ws3.cell(12, 1, "10초이내")
    ws3.cell(12, 2, hook.get("within_10sec", ""))
    ws3.cell(13, 1, "30초이내")
    ws3.cell(13, 2, hook.get("within_30sec", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────
# 메인 탭 렌더링
# ──────────────────────────────────────────

def render_thumbnail_tab():
    # 진행 상태 바
    render_pipeline_status()

    st.header("🎨 썸네일·제목 통합 전략 생성기")
    st.caption("프롬프트 1에서 선택한 주제를 바탕으로 CTR 5% 이상을 목표로 하는 썸네일 문구와 제목을 확정합니다.")

    # ── 1단계 확정 내용 확인 및 편집 ──
    p1_ready = render_p1_confirmed_card(editable=True)
    if not p1_ready:
        return

    # 확정된 값 읽기
    channel_name   = st.session_state.get(P1_CHANNEL, "")
    topic_title    = st.session_state.get(P1_TOPIC_TITLE, "")
    core_message   = st.session_state.get(P1_CORE_MESSAGE, "")
    target_emotion = st.session_state.get(P1_EMOTION, "")
    hook_sentence  = st.session_state.get(P1_HOOK, "")

    st.divider()

    # ── 추가 요구사항 입력 ──
    with st.expander("⚙️ 추가 요구사항 (선택)", expanded=False):
        extra_req = st.text_area(
            "이번 썸네일·제목 전략에 추가할 요구사항",
            placeholder="예) 공포감 극대화, 숫자 강조, 주인공 미등장 선호",
            height=80,
            key="p2_extra",
        )

    # ── 분석 실행 버튼 ──
    run_btn = st.button(
        "🎨 썸네일·제목 전략 생성",
        type="primary",
        use_container_width=True,
        disabled=not (channel_name and topic_title),
    )

    if run_btn:
        with st.spinner("Claude AI가 썸네일·제목 전략을 분석하는 중... (10~20초 소요)"):
            try:
                extra = st.session_state.get("p2_extra", "")
                result = call_claude_prompt2(
                    channel_name, topic_title, core_message,
                    target_emotion, hook_sentence, extra,
                )
                st.session_state[P2_RESULT] = result
                st.success("✅ 분석 완료!")

            except ValueError as e:
                st.error(str(e))
                st.info("💡 팁: 추가 요구사항을 비워두고 재시도해보세요.")
                return
            except RuntimeError as e:
                st.error(f"API 오류: {str(e)}")
                return
            except Exception as e:
                st.error(f"예기치 않은 오류: {str(e)}")
                return

    result = st.session_state.get(P2_RESULT)
    if not result:
        st.info("위 버튼을 눌러 썸네일·제목 전략을 생성하세요.")
        return

    st.divider()

    # ── 클릭 구조 진단 ──
    click = result.get("click_structure", {})
    with st.expander("🔍 클릭 구조 진단", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"**핵심 사건:** {click.get('core_event','')}")
            st.markdown(f"**핵심 결과:** {click.get('core_result','')}")
        with c2:
            st.markdown(f"**장면 우선도:** {click.get('scene_priority','')}")
            st.markdown(f"**주인공 등장:** {click.get('protagonist_needed','')}")

    st.divider()

    # ── 썸네일 문구 5종 ──
    st.subheader("💬 썸네일 문구 5종")
    st.caption("선택할 문구를 고르고 '이 문구로 확정' 버튼을 누르세요.")

    thumbnails = result.get("thumbnails", [])
    selected_thumb_id = st.session_state.get("p2_selected_thumb_id", 0)

    for thumb in thumbnails:
        render_thumbnail_card(thumb, selected_thumb_id)
        if st.button(
            f"✅ 문구 {thumb.get('id')} 확정",
            key=f"select_thumb_{thumb.get('id')}",
        ):
            st.session_state["p2_selected_thumb_id"] = thumb.get("id")
            thumb_text = (
                f"[말풍선] {thumb.get('speech_bubble','')} ({thumb.get('speech_bubble_color','')})\n"
                f"[1행] {thumb.get('line1','')} ({thumb.get('line1_color','')})\n"
                f"[2행] {thumb.get('line2','')} ({thumb.get('line2_color','')})"
            )
            st.session_state[P2_THUMBNAIL] = thumb_text
            st.success(f"문구 {thumb.get('id')} 확정!")
            st.rerun()

    st.divider()

    # ── 제목 5종 ──
    st.subheader("📝 제목 5종")
    st.caption("확정할 제목을 선택하세요.")

    titles = result.get("titles", [])
    selected_title_id = st.session_state.get("p2_selected_title_id", 0)

    for title in titles:
        tid = title.get("id", 0)
        is_selected = (tid == selected_title_id)
        border = "2px solid #4A90E2" if is_selected else "1px solid #e0e0e0"
        bg = "#f0f7ff" if is_selected else "#ffffff"
        badge = "✅ 선택됨" if is_selected else ""

        st.markdown(
            f"""
            <div style="border:{border}; border-radius:10px; padding:12px;
                        margin-bottom:8px; background:{bg};">
                <div style="display:flex; justify-content:space-between;">
                    <span style="font-size:12px; color:#888;">제목 {tid}</span>
                    <span style="font-size:12px; color:#4A90E2; font-weight:600;">{badge}</span>
                </div>
                <div style="font-size:15px; font-weight:600; margin:6px 0;">
                    {title.get('title','')}
                </div>
                <div style="display:flex; gap:8px; flex-wrap:wrap;">
                    <span style="font-size:11px; background:#eef; padding:2px 8px; border-radius:20px;">
                        키워드: {title.get('main_keyword','')}
                    </span>
                    <span style="font-size:11px; background:#eef; padding:2px 8px; border-radius:20px;">
                        감정: {title.get('emotion_device','')}
                    </span>
                    <span style="font-size:11px; background:#eef; padding:2px 8px; border-radius:20px;">
                        검색적합: {title.get('search_fit','')}
                    </span>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button(f"✅ 제목 {tid} 확정", key=f"select_title_{tid}"):
            st.session_state["p2_selected_title_id"] = tid
            st.session_state[P2_TITLE] = title.get("title", "")
            st.success(f"제목 {tid} 확정!")
            st.rerun()

    st.divider()

    # ── Best 조합 추천 ──
    st.subheader("🔗 썸네일-제목 Best 조합 3")
    for combo in result.get("best_combinations", []):
        with st.expander(
            f"조합 {combo.get('rank')} — 썸네일 {combo.get('thumbnail_id')} + 제목 {combo.get('title_id')}"
        ):
            st.markdown(f"**선정 이유:** {combo.get('reason','')}")
            st.markdown(f"**감정 기능:** {combo.get('emotion_function','')}")
            st.markdown(f"**검색 기능:** {combo.get('search_function','')}")
            st.markdown(f"**Hook 연결:** {combo.get('hook_connection','')}")

            if st.button(
                "이 조합으로 한 번에 확정",
                key=f"combo_{combo.get('rank')}",
            ):
                thumb_id = combo.get("thumbnail_id")
                title_id = combo.get("title_id")
                thumb = next((t for t in thumbnails if t.get("id") == thumb_id), {})
                title_obj = next((t for t in titles if t.get("id") == title_id), {})

                st.session_state["p2_selected_thumb_id"] = thumb_id
                st.session_state["p2_selected_title_id"] = title_id
                thumb_text = (
                    f"[말풍선] {thumb.get('speech_bubble','')} ({thumb.get('speech_bubble_color','')})\n"
                    f"[1행] {thumb.get('line1','')} ({thumb.get('line1_color','')})\n"
                    f"[2행] {thumb.get('line2','')} ({thumb.get('line2_color','')})"
                )
                st.session_state[P2_THUMBNAIL] = thumb_text
                st.session_state[P2_TITLE] = title_obj.get("title", "")
                st.success(f"조합 {combo.get('rank')} 확정 완료!")
                st.rerun()

    st.divider()

    # ── 초반 30초 Hook ──
    st.subheader("🎬 초반 30초 Hook 전략")
    hook = result.get("hook_30sec", {})
    st.markdown(f"**첫 문장:** {hook.get('first_sentence','')}")
    st.markdown(f"**10초 이내:** {hook.get('within_10sec','')}")
    st.markdown(f"**30초 이내:** {hook.get('within_30sec','')}")

    if st.button("✅ 이 Hook 전략 확정", key="confirm_hook"):
        hook_text = (
            f"첫문장: {hook.get('first_sentence','')}\n"
            f"10초이내: {hook.get('within_10sec','')}\n"
            f"30초이내: {hook.get('within_30sec','')}"
        )
        st.session_state[P2_HOOK_30SEC] = hook_text
        st.success("Hook 전략이 저장되었습니다!")

    st.divider()

    # ── 이미지 프롬프트 3종 ──
    st.subheader("🖼️ 썸네일 이미지 프롬프트 3종 (나노바나나 PRO용)")
    for img in result.get("image_prompts", []):
        with st.expander(f"프롬프트 {img.get('id')} — {img.get('concept','')}"):
            prompt_text = img.get("prompt_en", "")
            st.text_area(
                "영어 프롬프트 (복사하여 사용)",
                value=prompt_text,
                height=200,
                key=f"img_prompt_{img.get('id')}",
            )
            if st.button(f"📋 이 프롬프트 선택 저장", key=f"save_img_{img.get('id')}"):
                st.session_state[P2_IMAGE_PROMPT] = prompt_text
                st.success("이미지 프롬프트가 저장되었습니다!")

    st.divider()

    # ── 확정 내용 요약 + 다음 단계 안내 ──
    thumb_confirmed = st.session_state.get(P2_THUMBNAIL, "")
    title_confirmed = st.session_state.get(P2_TITLE, "")

    if thumb_confirmed and title_confirmed:
        st.success(
            f"✅ **2단계 확정 완료!**\n\n"
            f"📌 확정 제목: {title_confirmed}\n"
            f"🖼️ 확정 썸네일:\n{thumb_confirmed}"
        )
        st.info("👉 다음 단계인 **'📐 대본 구조 설계'** 탭으로 이동하세요. (향후 추가 예정)")
    else:
        missing = []
        if not thumb_confirmed:
            missing.append("썸네일 문구")
        if not title_confirmed:
            missing.append("제목")
        st.warning(f"⚠️ 아직 확정되지 않은 항목: {', '.join(missing)}")

    # ── 내보내기 ──
    st.divider()
    ec1, ec2 = st.columns(2)
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    with ec1:
        excel_bytes = export_p2_excel(result, channel_name, topic_title)
        st.download_button(
            "📥 Excel 다운로드",
            data=excel_bytes,
            file_name=f"썸네일전략_{channel_name}_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with ec2:
        hook_confirmed = st.session_state.get(P2_HOOK_30SEC, "")
        summary = (
            f"채널,{channel_name}\n"
            f"주제,{topic_title}\n"
            f"제목,\"{title_confirmed}\"\n"
            f"썸네일,\"{thumb_confirmed}\"\n"
            f"Hook,\"{hook_confirmed}\"\n"
        )
        st.download_button(
            "📥 확정 내용 CSV",
            data=summary.encode("utf-8-sig"),
            file_name=f"확정내용_{channel_name}_{ts}.csv",
            mime="text/csv",
            use_container_width=True,
        )
