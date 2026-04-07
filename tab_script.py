import re
import json
import io
import os
from datetime import datetime

import streamlit as st
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from channel_db import CHANNEL_DB
from prompts import PROMPT_4_SYSTEM, PROMPT_4_SECTION_SYSTEM
from session_state_manager import (
    P1_CHANNEL, P1_TOPIC_TITLE, P1_CORE_MESSAGE, P1_EMOTION, P1_HOOK,
    P2_TITLE, P2_THUMBNAIL, P2_HOOK_30SEC,
    P3_RESULT, P3_VIDEO_LENGTH, P3_STRUCTURE, P3_EMOTION_MAP, P3_MINI_HOOKS, P3_SCENE_META,
    P4_RESULT, P4_SCRIPT, P4_TOTAL_WORDS, P4_WRITING_NOTES,
    render_pipeline_status, render_p1_confirmed_card, render_p2_confirmed_card,
)


# ──────────────────────────────────────────
# 상수
# ──────────────────────────────────────────

SECTION_CONFIG = {
    "hook":     {"label": "HOOK",     "timecode": "00:00", "color": "#E53935", "target_words": 500},
    "teaser":   {"label": "TEASER",   "timecode": "01:00", "color": "#FB8C00", "target_words": 500},
    "big_idea": {"label": "BIG IDEA", "timecode": "02:00", "color": "#F9A825", "target_words": 500},
    "intro":    {"label": "INTRO",    "timecode": "03:00", "color": "#43A047", "target_words": 500},
    "body1":    {"label": "BODY 1",   "timecode": "04:00", "color": "#00897B", "target_words": 1600},
    "body2":    {"label": "BODY 2",   "timecode": "07:00", "color": "#039BE5", "target_words": 1600},
    "body3":    {"label": "BODY 3",   "timecode": "10:15", "color": "#1E88E5", "target_words": 1600},
    "body4":    {"label": "BODY 4",   "timecode": "13:30", "color": "#5E35B1", "target_words": 1600},
    "reveal":   {"label": "REVEAL",   "timecode": "17:00", "color": "#8E24AA", "target_words": 750},
    "impact":   {"label": "IMPACT",   "timecode": "18:30", "color": "#D81B60", "target_words": 250},
    "end":      {"label": "END",      "timecode": "19:00", "color": "#546E7A", "target_words": 500},
}

SECTION_ORDER = ["hook", "teaser", "big_idea", "intro", "body1", "body2", "body3", "body4", "reveal", "impact", "end"]

SECTION_SPECIAL = {
    "hook": "첫 문장에서 핵심 약속 즉시 등장. 클릭 이유를 3초 안에 제시.",
    "teaser": "영상 전체 예고. 시청자가 기대할 3가지 포인트 명시.",
    "big_idea": "핵심 아이디어를 숫자/사물/대비로 시각화. 추상적 설명 금지.",
    "intro": "주인공 소개 + 문제 상황 공감. 시청자가 '나 얘기네'라고 느끼도록.",
    "body1": "첫 번째 핵심 증거. 구체적 사례 또는 데이터 중심. 말미에 [MINI-HOOK].",
    "body2": "두 번째 핵심 증거. 대조 인물 또는 군중 반응 장면 포함. 말미에 [MINI-HOOK].",
    "body3": "세 번째 핵심 증거. 결과·여파 장면 중심. 감정 급변 포인트 필수. 말미에 [MINI-HOOK].",
    "body4": "네 번째 핵심 증거. 클라이맥스 직전 최대 긴장감 조성. 말미에 [MINI-HOOK].",
    "reveal": "가장 강렬한 반전 또는 핵심 진실. 시청자가 '아!'라고 외칠 순간.",
    "impact": "시청자 삶과 직접 연결. '당신도 지금 이 상황일 수 있다'는 공감.",
    "end": "명확한 행동 촉구 + 감성적 여운. 다음 영상 예고 또는 구독 CTA 자연스럽게.",
}


# ──────────────────────────────────────────
# 클라이언트
# ──────────────────────────────────────────

def _get_client():
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        key = ""
    key = key or os.environ.get("ANTHROPIC_API_KEY", "")
    return anthropic.Anthropic(api_key=key)


# ──────────────────────────────────────────
# JSON 파싱 방어
# ──────────────────────────────────────────

def _extract_json(text: str) -> str:
    text = text.strip()
    m = re.search(r"```json\s*([\s\S]*?)\s*```", text)
    if m:
        return m.group(1).strip()
    m = re.search(r"```\s*([\s\S]*?)\s*```", text)
    if m:
        candidate = m.group(1).strip()
        if candidate.startswith("{"):
            return candidate
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return text[start:end + 1]
    return text


def _safe_loads(text: str) -> dict:
    json_str = _extract_json(text)
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        preview = json_str[:300].replace("\n", " ")
        raise json.JSONDecodeError(
            f"JSON 파싱 실패 (앞 300자: {preview}...)", e.doc, e.pos
        )


# ──────────────────────────────────────────
# 페르소나 블록
# ──────────────────────────────────────────

def build_persona_block(channel_name: str) -> str:
    if channel_name not in CHANNEL_DB:
        return "채널 정보 없음 — 중립적 분석가 페르소나 적용"
    info = CHANNEL_DB[channel_name]
    return (
        f"채널명: {channel_name}\n"
        f"주인공: {info['host']} ({info['host_desc']})\n"
        f"색상: {info['color_primary']} / {info['color_secondary']}\n"
        f"톤앤매너: {info['tone']}\n"
        f"타겟: {info['target']}\n"
        f"썸네일 전략: {info['thumbnail_style']}\n"
        f"시각 무드: {info['visual_mood']}"
    )


# ──────────────────────────────────────────
# 구조 요약 텍스트 빌드
# ──────────────────────────────────────────

def build_structure_summary(structure: dict) -> str:
    lines = []
    for key in SECTION_ORDER:
        sec = structure.get(key, {})
        cfg = SECTION_CONFIG.get(key, {})
        label = cfg.get("label", key.upper())
        tc = cfg.get("timecode", "")
        info = sec.get("info_purpose", "")
        emotion = sec.get("emotion_goal", "")
        scene = sec.get("scene_type", "")
        role = sec.get("protagonist_role", "")
        if info or emotion:
            line = f"[{tc}] {label}: 정보={info} / 감정={emotion}"
            if scene:
                line += f" / 장면={scene}"
            if role:
                line += f" / 주인공={role}"
            lines.append(line)
    return "\n".join(lines)


# ──────────────────────────────────────────
# Claude API: 전체 대본 작성
# ──────────────────────────────────────────

def call_claude_prompt4_full(
    channel_name: str,
    topic_title: str,
    core_message: str,
    target_emotion: str,
    confirmed_title: str,
    confirmed_thumbnail: str,
    hook_30sec: str,
    video_length: str,
    structure: dict,
    scene_meta: dict,
) -> dict:
    persona_block = build_persona_block(channel_name)
    structure_summary = build_structure_summary(structure)
    prompt4_instruction = scene_meta.get(
        "prompt4_instruction",
        "주인공이 해석하고 장면이 증명하는 구조. 각 2~4문장마다 시각 큐 삽입."
    )

    system_prompt = PROMPT_4_SYSTEM.format(
        persona_block=persona_block,
        channel_name=channel_name,
        topic_title=topic_title,
        core_message=core_message,
        target_emotion=target_emotion,
        confirmed_title=confirmed_title,
        confirmed_thumbnail=confirmed_thumbnail,
        hook_30sec=hook_30sec,
        video_length=video_length,
        structure_summary=structure_summary,
        prompt4_instruction=prompt4_instruction,
    )

    user_message = (
        f"채널: {channel_name} / 주제: {topic_title}\n"
        f"제목: {confirmed_title}\n"
        f"영상 길이: {video_length}\n\n"
        "위 구조대로 전체 대본을 작성하고 JSON만 반환하라.\n"
        "응답은 반드시 { 로 시작하고 } 로 끝나야 한다."
    )

    MAX_ATTEMPTS = 3
    last_raw = ""
    last_error = None
    client = _get_client()

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=16000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_message}],
            )
            last_raw = response.content[0].text
            return _safe_loads(last_raw)

        except json.JSONDecodeError as e:
            last_error = e
            if attempt < MAX_ATTEMPTS:
                user_message = (
                    f"이전 응답이 유효한 JSON이 아니었다.\n오류: {str(e)}\n"
                    f"이전 응답 앞부분: {last_raw[:300]}\n\n"
                    "규칙 재확인 후 올바른 JSON만 반환하라:\n"
                    "1. 응답 첫 글자는 반드시 { 이어야 한다\n"
                    "2. 마크다운 코드블록(```) 절대 사용 금지\n"
                    "3. 설명 텍스트 절대 금지\n\n"
                    f"주제: {topic_title} / JSON만 반환하라."
                )
        except Exception as e:
            raise RuntimeError(f"API 호출 오류: {str(e)}")

    raise ValueError(
        f"Claude API가 {MAX_ATTEMPTS}회 시도 후에도 유효한 JSON을 반환하지 못했습니다.\n"
        f"마지막 오류: {str(last_error)}"
    )


# ──────────────────────────────────────────
# Claude API: 섹션 개별 재작성
# ──────────────────────────────────────────

def call_claude_prompt4_section(
    section_key: str,
    channel_name: str,
    topic_title: str,
    core_message: str,
    target_emotion: str,
    confirmed_title: str,
    structure: dict,
    scene_meta: dict,
) -> dict:
    persona_block = build_persona_block(channel_name)
    sec = structure.get(section_key, {})
    cfg = SECTION_CONFIG.get(section_key, {})
    prompt4_instruction = scene_meta.get(
        "prompt4_instruction",
        "주인공이 해석하고 장면이 증명하는 구조."
    )

    system_prompt = PROMPT_4_SECTION_SYSTEM.format(
        persona_block=persona_block,
        channel_name=channel_name,
        topic_title=topic_title,
        core_message=core_message,
        target_emotion=target_emotion,
        confirmed_title=confirmed_title,
        section_label=cfg.get("label", section_key.upper()),
        timecode=cfg.get("timecode", ""),
        word_count_target=cfg.get("target_words", 500),
        info_purpose=sec.get("info_purpose", ""),
        emotion_goal=sec.get("emotion_goal", ""),
        scene_type=sec.get("scene_type", ""),
        protagonist_role=sec.get("protagonist_role", ""),
        supporting_characters=sec.get("supporting_characters", ""),
        key_objects=sec.get("key_objects", ""),
        special_instruction=SECTION_SPECIAL.get(section_key, ""),
        section_key=section_key,
    )

    user_message = (
        f"[{cfg.get('timecode','')}] {cfg.get('label', section_key.upper())} 섹션 대본을 작성하고 JSON만 반환하라.\n"
        "응답은 반드시 { 로 시작하고 } 로 끝나야 한다."
    )

    MAX_ATTEMPTS = 3
    last_raw = ""
    last_error = None
    client = _get_client()

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=4000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_message}],
            )
            last_raw = response.content[0].text
            return _safe_loads(last_raw)

        except json.JSONDecodeError as e:
            last_error = e
            if attempt < MAX_ATTEMPTS:
                user_message = (
                    f"이전 응답이 유효한 JSON이 아니었다.\n오류: {str(e)}\n"
                    "규칙 재확인 후 올바른 JSON만 반환하라: "
                    "응답 첫 글자 {{ / 코드블록 금지 / JSON만 반환."
                )
        except Exception as e:
            raise RuntimeError(f"API 호출 오류: {str(e)}")

    raise ValueError(f"섹션 작성 실패 ({section_key}): {str(last_error)}")


# ──────────────────────────────────────────
# 대본 텍스트 렌더링 (색상 태그 하이라이트)
# ──────────────────────────────────────────

def render_script_text(text: str, section_color: str):
    """대본 텍스트를 시각 큐 하이라이트와 함께 표시한다."""
    # [SCENE:...], [CUT:...], [B-ROLL:...], [TEXT:...], [MINI-HOOK] 하이라이트
    def highlight(m):
        tag = m.group(1)
        content = m.group(2).strip() if m.group(2) else ""
        tag_colors = {
            "SCENE": "#e3f2fd", "CUT": "#f3e5f5", "B-ROLL": "#e8f5e9",
            "TEXT": "#fff8e1", "MINI-HOOK": "#fce4ec",
        }
        bg = tag_colors.get(tag, "#f5f5f5")
        if content:
            return (
                f'<span style="background:{bg}; color:#333; border-radius:4px;'
                f' padding:1px 6px; font-size:11px; font-weight:600;">'
                f'[{tag}: {content}]</span>'
            )
        return (
            f'<span style="background:{bg}; color:#333; border-radius:4px;'
            f' padding:1px 6px; font-size:12px; font-weight:700;">'
            f'[{tag}]</span>'
        )

    # 시각 큐 처리
    display = re.sub(
        r'\[(SCENE|CUT|B-ROLL|TEXT|MINI-HOOK)(?::\s*)?(.*?)\]',
        highlight,
        text,
        flags=re.IGNORECASE
    )
    # 줄바꿈을 <br>로 변환
    display = display.replace("\n", "<br>")

    st.markdown(
        f"""
        <div style="
            border-left: 4px solid {section_color};
            background: #fafafa;
            border-radius: 6px;
            padding: 14px 16px;
            font-size: 14px;
            line-height: 1.8;
            color: #1a1a1a;
            white-space: pre-wrap;
        ">{display}</div>
        """,
        unsafe_allow_html=True,
    )


# ──────────────────────────────────────────
# Excel 내보내기
# ──────────────────────────────────────────

def _header_style(cell, bg="4A90E2"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _data_style(cell, bg="FFFFFF"):
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.fill = PatternFill("solid", fgColor=bg)


def export_p4_excel(script: dict, topic_title: str, channel_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "✍️ 대본"

    # 헤더 정보
    for i, (k, v) in enumerate([("채널명", channel_name), ("주제", topic_title),
                                  ("작성 일시", datetime.now().strftime("%Y-%m-%d %H:%M"))], 1):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    ws.append([])

    headers = ["타임코드", "섹션", "목표 글자수", "실제 글자수", "대본 내용"]
    row_start = 5
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        _header_style(ws.cell(row_start, col_idx))
    ws.row_dimensions[row_start].height = 30

    total_words = 0
    for idx, key in enumerate(SECTION_ORDER):
        sec = script.get(key, {})
        cfg = SECTION_CONFIG.get(key, {})
        text = sec.get("text", "")
        wc = sec.get("word_count", len(text))
        total_words += wc
        bg = "F8F9FA" if idx % 2 == 0 else "FFFFFF"
        row = [
            cfg.get("timecode", ""),
            cfg.get("label", key.upper()),
            cfg.get("target_words", 0),
            wc,
            text,
        ]
        ws.append(row)
        for col_idx, _ in enumerate(row, 1):
            _data_style(ws.cell(ws.max_row, col_idx), bg)
        ws.row_dimensions[ws.max_row].height = max(60, min(len(text) // 3, 200))

    # 합계 행
    ws.append(["", "합계", "", total_words, ""])
    for col_idx in range(1, 6):
        ws.cell(ws.max_row, col_idx).font = Font(bold=True)
        ws.cell(ws.max_row, col_idx).fill = PatternFill("solid", fgColor="E8F0FE")

    for i, w in enumerate([12, 14, 14, 14, 80], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_p4_txt(script: dict, topic_title: str) -> bytes:
    lines = [f"# {topic_title}", f"작성 일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}", "=" * 60, ""]
    for key in SECTION_ORDER:
        sec = script.get(key, {})
        cfg = SECTION_CONFIG.get(key, {})
        label = cfg.get("label", key.upper())
        tc = cfg.get("timecode", "")
        text = sec.get("text", "")
        wc = sec.get("word_count", len(text))
        lines.append(f"[{tc}] {label}  ({wc}자)")
        lines.append("-" * 40)
        lines.append(text)
        lines.append("")
    return "\n".join(lines).encode("utf-8")


# ──────────────────────────────────────────
# 메인 탭 렌더링
# ──────────────────────────────────────────

def render_script_tab():
    render_pipeline_status()

    st.header("✍️ 대본 작성", divider="gray")
    st.caption("대본 구조 설계를 바탕으로 실제 촬영·편집에 사용할 완성 대본을 작성합니다.")

    # ── 이전 단계 카드 ───────────────────────────────────────────────────────
    ok1 = render_p1_confirmed_card(editable=False)
    if not ok1:
        return

    ok2 = render_p2_confirmed_card(editable=False)
    if not ok2:
        return

    # P3 구조 확인
    structure = st.session_state.get(P3_STRUCTURE)
    if not structure:
        st.warning("⚠️ 프롬프트 3(대본 구조 설계)를 먼저 완료해주세요.")
        return

    with st.expander("📌 3단계 구조 요약 확인", expanded=False):
        scene_meta = st.session_state.get(P3_SCENE_META, {})
        video_length = st.session_state.get(P3_VIDEO_LENGTH, "20분 내외")
        st.markdown(f"**영상 길이:** {video_length}")
        for key in SECTION_ORDER:
            sec = structure.get(key, {})
            cfg = SECTION_CONFIG.get(key, {})
            info = sec.get("info_purpose", "")
            if info:
                st.caption(f"[{cfg.get('timecode','')}] **{cfg.get('label','')}**: {info}")

    st.divider()

    # ── 작성 모드 선택 ────────────────────────────────────────────────────────
    col_mode, col_btn = st.columns([3, 1])
    with col_mode:
        write_mode = st.radio(
            "작성 방식",
            options=["전체 대본 한번에", "섹션 선택 작성"],
            horizontal=True,
            key="p4_write_mode",
        )

    selected_sections = SECTION_ORDER
    if write_mode == "섹션 선택 작성":
        section_options = {
            f"[{SECTION_CONFIG[k]['timecode']}] {SECTION_CONFIG[k]['label']}": k
            for k in SECTION_ORDER
        }
        chosen_labels = st.multiselect(
            "작성할 섹션 선택",
            options=list(section_options.keys()),
            default=list(section_options.keys()),
            key="p4_section_select",
        )
        selected_sections = [section_options[l] for l in chosen_labels]

    run_btn = st.button(
        "✍️ 대본 작성 시작" if write_mode == "전체 대본 한번에" else "✍️ 선택 섹션 작성",
        key="p4_run_btn",
        type="primary",
        use_container_width=True,
    )

    # ── API 호출 ───────────────────────────────────────────────────────────────
    channel_name = st.session_state.get(P1_CHANNEL, "")
    topic_title = st.session_state.get(P1_TOPIC_TITLE, "")
    core_message = st.session_state.get(P1_CORE_MESSAGE, "")
    target_emotion = st.session_state.get(P1_EMOTION, "")
    confirmed_title = st.session_state.get(P2_TITLE, "")
    confirmed_thumbnail = st.session_state.get(P2_THUMBNAIL, "")
    hook_30sec = st.session_state.get(P2_HOOK_30SEC, "")
    video_length = st.session_state.get(P3_VIDEO_LENGTH, "20분 내외")
    scene_meta = st.session_state.get(P3_SCENE_META, {})

    if run_btn:
        if write_mode == "전체 대본 한번에":
            with st.spinner("Claude AI가 전체 대본을 작성 중입니다... (30-60초 소요)"):
                try:
                    result = call_claude_prompt4_full(
                        channel_name=channel_name,
                        topic_title=topic_title,
                        core_message=core_message,
                        target_emotion=target_emotion,
                        confirmed_title=confirmed_title,
                        confirmed_thumbnail=confirmed_thumbnail,
                        hook_30sec=hook_30sec,
                        video_length=video_length,
                        structure=structure,
                        scene_meta=scene_meta,
                    )
                    script = result.get("script", {})
                    # 기존 스크립트와 병합 (새로 작성된 섹션만 업데이트)
                    existing = st.session_state.get(P4_SCRIPT, {})
                    existing.update(script)
                    st.session_state[P4_RESULT] = result
                    st.session_state[P4_SCRIPT] = existing
                    st.session_state[P4_TOTAL_WORDS] = result.get("total_word_count", 0)
                    st.session_state[P4_WRITING_NOTES] = result.get("writing_notes", "")
                    st.success("전체 대본 작성 완료!")
                except ValueError as e:
                    st.error(f"대본 작성 오류: {e}")
                except RuntimeError as e:
                    st.error(f"API 오류: {e}")
                except Exception as e:
                    st.error(f"오류 발생: {e}")
        else:
            # 섹션별 작성
            existing = st.session_state.get(P4_SCRIPT, {})
            progress_bar = st.progress(0)
            status_text = st.empty()
            failed = []
            for i, key in enumerate(selected_sections):
                cfg = SECTION_CONFIG.get(key, {})
                status_text.text(f"[{cfg.get('timecode','')}] {cfg.get('label','')} 작성 중...")
                try:
                    result = call_claude_prompt4_section(
                        section_key=key,
                        channel_name=channel_name,
                        topic_title=topic_title,
                        core_message=core_message,
                        target_emotion=target_emotion,
                        confirmed_title=confirmed_title,
                        structure=structure,
                        scene_meta=scene_meta,
                    )
                    existing[key] = result
                except Exception as e:
                    failed.append(f"{cfg.get('label', key)}: {e}")
                progress_bar.progress((i + 1) / len(selected_sections))

            st.session_state[P4_SCRIPT] = existing
            st.session_state[P4_RESULT] = {"script": existing}
            total = sum(s.get("word_count", 0) for s in existing.values())
            st.session_state[P4_TOTAL_WORDS] = total
            status_text.empty()
            progress_bar.empty()
            if failed:
                st.warning(f"일부 섹션 실패: {', '.join(failed)}")
            else:
                st.success(f"선택 섹션 대본 작성 완료! (총 {len(selected_sections)}개)")

    script = st.session_state.get(P4_SCRIPT, {})
    if not script:
        st.info("'대본 작성 시작' 버튼을 눌러 대본을 생성하세요.", icon="✍️")
        return

    # ── 통계 메트릭 ────────────────────────────────────────────────────────────
    st.subheader("📊 대본 통계")
    total_words = sum(s.get("word_count", len(s.get("text", ""))) for s in script.values() if isinstance(s, dict))
    completed = sum(1 for k in SECTION_ORDER if script.get(k, {}).get("text", ""))
    avg_words = total_words // max(completed, 1)

    metric_cols = st.columns(4)
    metric_cols[0].metric("✍️ 완성 섹션", f"{completed} / {len(SECTION_ORDER)}")
    metric_cols[1].metric("📝 총 글자수", f"{total_words:,}자")
    metric_cols[2].metric("📏 평균 섹션 글자수", f"{avg_words:,}자")
    writing_notes = st.session_state.get(P4_WRITING_NOTES, "")
    if writing_notes:
        metric_cols[3].metric("📋 작성 노트", "확인 ▼")

    if writing_notes:
        with st.expander("📋 작성 노트 보기", expanded=False):
            st.write(writing_notes)

    st.divider()

    # ── 대본 표시 모드 ─────────────────────────────────────────────────────────
    view_mode = st.radio(
        "보기 모드",
        options=["섹션별 카드", "전체 대본 연속"],
        horizontal=True,
        key="p4_view_mode",
    )

    if view_mode == "섹션별 카드":
        for key in SECTION_ORDER:
            sec = script.get(key, {})
            if not isinstance(sec, dict):
                continue
            text = sec.get("text", "")
            cfg = SECTION_CONFIG.get(key, {})
            label = cfg.get("label", key.upper())
            timecode = cfg.get("timecode", "")
            color = cfg.get("color", "#888")
            target = cfg.get("target_words", 0)
            wc = sec.get("word_count", len(text))

            # 섹션 헤더
            col_hdr, col_meta, col_rewrite = st.columns([3, 2, 1])
            with col_hdr:
                st.markdown(
                    f'<span style="color:{color}; font-size:16px; font-weight:700;">'
                    f'[{timecode}] {label}</span>',
                    unsafe_allow_html=True,
                )
            with col_meta:
                diff = wc - target
                diff_str = f"+{diff}" if diff >= 0 else str(diff)
                color_diff = "#4CAF50" if abs(diff) <= 100 else "#F44336"
                st.markdown(
                    f'<span style="font-size:12px; color:#888;">{wc}자 / 목표 {target}자 </span>'
                    f'<span style="font-size:12px; color:{color_diff};">({diff_str})</span>',
                    unsafe_allow_html=True,
                )
            with col_rewrite:
                if st.button("🔄 재작성", key=f"rewrite_{key}", help=f"{label} 섹션 재작성"):
                    with st.spinner(f"{label} 재작성 중..."):
                        try:
                            result = call_claude_prompt4_section(
                                section_key=key,
                                channel_name=channel_name,
                                topic_title=topic_title,
                                core_message=core_message,
                                target_emotion=target_emotion,
                                confirmed_title=confirmed_title,
                                structure=structure,
                                scene_meta=scene_meta,
                            )
                            script[key] = result
                            st.session_state[P4_SCRIPT] = script
                            st.rerun()
                        except Exception as e:
                            st.error(f"재작성 실패: {e}")

            if text:
                render_script_text(text, color)
            else:
                st.info(f"{label} 섹션이 아직 작성되지 않았습니다.")

            st.markdown("---")

    else:
        # 전체 대본 연속 보기
        full_lines = []
        for key in SECTION_ORDER:
            sec = script.get(key, {})
            if not isinstance(sec, dict):
                continue
            text = sec.get("text", "")
            cfg = SECTION_CONFIG.get(key, {})
            if text:
                full_lines.append(f"[{cfg.get('timecode','')}] {cfg.get('label','')} ({'─'*40})")
                full_lines.append(text)
                full_lines.append("")
        if full_lines:
            st.text_area(
                "전체 대본",
                value="\n".join(full_lines),
                height=800,
                key="p4_full_script_view",
            )

    # ── 내보내기 ────────────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📥 내보내기")

    export_cols = st.columns(2)
    with export_cols[0]:
        try:
            excel_bytes = export_p4_excel(
                script=script,
                topic_title=topic_title,
                channel_name=channel_name,
            )
            fname_xlsx = f"대본_{topic_title[:20]}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="📊 Excel 다운로드",
                data=excel_bytes,
                file_name=fname_xlsx,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="p4_download_excel",
                use_container_width=True,
            )
        except Exception as e:
            st.warning(f"Excel 생성 오류: {e}")

    with export_cols[1]:
        try:
            txt_bytes = export_p4_txt(script=script, topic_title=topic_title)
            fname_txt = f"대본_{topic_title[:20]}_{datetime.now().strftime('%Y%m%d_%H%M')}.txt"
            st.download_button(
                label="📄 TXT 다운로드",
                data=txt_bytes,
                file_name=fname_txt,
                mime="text/plain; charset=utf-8",
                key="p4_download_txt",
                use_container_width=True,
            )
        except Exception as e:
            st.warning(f"TXT 생성 오류: {e}")
